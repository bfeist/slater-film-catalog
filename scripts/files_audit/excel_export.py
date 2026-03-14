"""
Excel Export — Disk File Catalog for Expert Review.

Produces an .xlsx workbook where each sheet covers one top-level subfolder on
the O: drive.  Pre-filled columns carry everything the database already knows
about a file; highlighted yellow columns are left blank for the archive expert
to fill in (reel identifier, human title, shot-list PDF reference, notes).

A separate "Summary" sheet gives per-folder counts so the expert can see at a
glance how much still needs attention.

Any rows that were previously annotated (via excel_import.py) are pre-populated
from the file_annotations table so incremental reviews don't lose prior work.

Usage (run from scripts/files_audit/)
-----
Export every file under O:/Master 1 (walks disk):
    uv run python excel_export.py --root "O:/Master 1" --out exports/expert_review.xlsx

Skip the disk walk and export only files already indexed in the DB (fast):
    uv run python excel_export.py --root "O:/Master 1" --from-db --out exports/expert_review.xlsx

Export ALL indexed roots without walking the disk:
    uv run python excel_export.py --from-db --out exports/expert_review.xlsx

Export only files that have no matched reel identifier yet:
    uv run python excel_export.py --root "O:/Master 1" --unmatched-only --out exports/gaps.xlsx
"""

from __future__ import annotations

import argparse
import os
import sqlite3
import time
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DB_PATH = "../../database/catalog.db"
DRIVE_ROOT = "O:/"

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

# (header, source-hint, col-width)
AUTO_COLS = [
    ("subfolder",         "disk",  22),
    ("filename",          "disk",  40),
    ("extension",         "disk",   8),
    ("size",              "disk",  10),
    ("full_path",         "disk",  60),
    ("in_db",             "DB",     6),
    ("db_id",             "DB",     7),
    ("matched_reels",     "DB",    20),
    ("reel_title",        "DB",    55),
    ("has_shotlist_pdf",  "DB",     8),
]

EXPERT_COLS = [
    ("expert_identifier", "FILL IN — reel/roll number (e.g. FR-1234)", 20),
    ("expert_title",      "FILL IN — human title for this file",        50),
    ("shotlist_pdf",      "FILL IN — shot-list PDF filename if known",  30),
    ("notes",             "FILL IN — any other notes",                  40),
]

ALL_COLS = AUTO_COLS + EXPERT_COLS
NCOLS = len(ALL_COLS)

# Column indices (1-based) for key columns
COL_FULL_PATH     = 5   # E
COL_IN_DB         = 6   # F
COL_MATCHED_REELS = 8   # H
COL_EXPERT_START  = len(AUTO_COLS) + 1  # first expert column

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

FILL_HEADER_AUTO   = PatternFill("solid", fgColor="4472C4")   # blue
FILL_HEADER_EXPERT = PatternFill("solid", fgColor="C65911")   # dark orange
FILL_AUTO          = PatternFill("solid", fgColor="D9E1F2")   # light blue-grey
FILL_NEEDS_ATTN    = PatternFill("solid", fgColor="FFEB9C")   # light amber — unmatched
FILL_EXPERT_CELL   = PatternFill("solid", fgColor="FFF2CC")   # pale yellow
FILL_EXPERT_PREFIL = PatternFill("solid", fgColor="E2EFDA")   # light green (previously saved)

FONT_HEADER   = Font(bold=True, color="FFFFFF", size=10)
FONT_EXPERT_H = Font(bold=True, color="FFFFFF", size=10)
FONT_NORMAL   = Font(size=9)
FONT_PATH     = Font(size=8, color="595959")


def _fill(cell, fill, font=None, alignment=None):
    cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def fmt_bytes(n: int) -> str:
    for unit, threshold in [("TB", 1 << 40), ("GB", 1 << 30), ("MB", 1 << 20), ("KB", 1 << 10)]:
        if abs(n) >= threshold:
            return f"{n / threshold:.1f} {unit}"
    return f"{n} B"


def normalise(path: str) -> str:
    return path.replace("\\", "/").rstrip("/").lower()


def top_folder_label(full_path: str, root: str) -> str:
    """Return the immediate child folder name, or '' if the file is directly in root."""
    rel = os.path.relpath(full_path, root)
    parts = Path(rel).parts
    return parts[0] if len(parts) > 1 else ""


def safe_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no special chars."""
    for ch in r'\/:*?[]':
        name = name.replace(ch, "-")
    return name[:31]


# ---------------------------------------------------------------------------
# DB loading
# ---------------------------------------------------------------------------

def load_db_data(db_path: str, root_filter: str | None = None) -> dict:
    """
    Returns a dict with:
      files          -- {normalised_path: {db_id, filename, folder_root, rel_path,
                                           size_bytes, extension}}
      file_matches   -- {file_id: [reel_identifier, ...]}
      reel_titles    -- {identifier: title}
      reel_shotlists -- {identifier: has_shotlist_pdf}
      annotations    -- {normalised_path: {expert_identifier, expert_title,
                                           shotlist_pdf, notes}}
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    # files_on_disk
    if root_filter:
        # Match any folder_root that starts with the normalised root prefix
        norm_root = root_filter.replace("\\", "/").rstrip("/")
        rows = conn.execute(
            "SELECT * FROM files_on_disk WHERE folder_root LIKE ?",
            (norm_root + "%",),
        ).fetchall()
        # Also try exact match for subfolders
        if not rows:
            rows = conn.execute("SELECT * FROM files_on_disk").fetchall()
            rows = [
                r for r in rows
                if normalise(r["folder_root"]).startswith(normalise(norm_root))
            ]
    else:
        rows = conn.execute("SELECT * FROM files_on_disk").fetchall()

    files: dict[str, dict] = {}
    for r in rows:
        full = (r["folder_root"].rstrip("/") + "/" + r["rel_path"]).replace("\\", "/")
        norm = normalise(full)
        files[norm] = {
            "db_id":       r["id"],
            "filename":    r["filename"],
            "folder_root": r["folder_root"],
            "rel_path":    r["rel_path"],
            "size_bytes":  r["size_bytes"] or 0,
            "extension":   r["extension"] or "",
            "full_path":   full,
        }

    # transfer_file_matches — map file_id -> list of reel identifiers
    match_rows = conn.execute(
        "SELECT file_id, reel_identifier FROM transfer_file_matches WHERE reel_identifier IS NOT NULL"
    ).fetchall()
    file_matches: dict[int, list[str]] = defaultdict(list)
    for mr in match_rows:
        if mr["reel_identifier"] not in file_matches[mr["file_id"]]:
            file_matches[mr["file_id"]].append(mr["reel_identifier"])

    # film_rolls — titles and shotlist flags
    roll_rows = conn.execute(
        "SELECT identifier, title, has_shotlist_pdf FROM film_rolls"
    ).fetchall()
    reel_titles:    dict[str, str]  = {r["identifier"]: r["title"] or "" for r in roll_rows}
    reel_shotlists: dict[str, int]  = {r["identifier"]: r["has_shotlist_pdf"] or 0 for r in roll_rows}

    # file_annotations (may not exist yet)
    annotations: dict[str, dict] = {}
    try:
        ann_rows = conn.execute(
            "SELECT full_path, expert_identifier, expert_title, shotlist_pdf, notes "
            "FROM file_annotations"
        ).fetchall()
        for ar in ann_rows:
            annotations[normalise(ar["full_path"])] = {
                "expert_identifier": ar["expert_identifier"] or "",
                "expert_title":      ar["expert_title"] or "",
                "shotlist_pdf":      ar["shotlist_pdf"] or "",
                "notes":             ar["notes"] or "",
            }
    except sqlite3.OperationalError:
        pass  # table doesn't exist yet — fine

    conn.close()
    print(
        f"  Loaded {len(files):,} indexed files | "
        f"{sum(len(v) for v in file_matches.values()):,} matches | "
        f"{len(annotations):,} existing annotations"
    )
    return {
        "files": files,
        "file_matches": dict(file_matches),
        "reel_titles": reel_titles,
        "reel_shotlists": reel_shotlists,
        "annotations": annotations,
    }


# ---------------------------------------------------------------------------
# Disk walk
# ---------------------------------------------------------------------------

def walk_root(root: str, db_files: dict[str, dict]) -> list[dict]:
    """
    Walk 'root' and return a list of file-info dicts.
    Files found on disk but not in the DB are included with in_db=False.
    """
    results: list[dict] = []
    scanned = 0
    t_last = time.time()
    root_norm = root.replace("\\", "/").rstrip("/")

    print(f"\n  Walking {root} ...")

    for dirpath, dirnames, filenames in os.walk(root, onerror=lambda e: None):
        dirnames.sort()
        for fname in sorted(filenames):
            full_path = os.path.join(dirpath, fname).replace("\\", "/")
            norm = normalise(full_path)
            db_rec = db_files.get(norm)

            if db_rec:
                size_bytes = db_rec["size_bytes"]
                extension  = db_rec["extension"]
                db_id      = db_rec["db_id"]
                in_db      = True
            else:
                try:
                    size_bytes = os.path.getsize(full_path)
                except OSError:
                    size_bytes = 0
                extension = Path(fname).suffix.lower()
                db_id     = None
                in_db     = False

            subfolder = top_folder_label(full_path, root_norm)

            results.append({
                "subfolder":  subfolder,
                "filename":   fname,
                "extension":  extension,
                "size_bytes": size_bytes,
                "full_path":  full_path,
                "norm_path":  norm,
                "in_db":      in_db,
                "db_id":      db_id,
            })
            scanned += 1
            if time.time() - t_last > 30:
                t_last = time.time()
                print(f"    ... {scanned:,} files  [{dirpath[:80]}]")

    print(f"  Walk complete: {scanned:,} files")
    return results


def files_from_db(db_files: dict[str, dict], root_filter: str | None = None) -> list[dict]:
    """Build the file list from DB records only (no disk walk)."""
    results: list[dict] = []
    norm_root = normalise(root_filter) if root_filter else None

    for norm_path, rec in db_files.items():
        if norm_root and not norm_path.startswith(norm_root):
            continue
        full_path = rec["full_path"]
        root_of   = rec["folder_root"].replace("\\", "/")
        # Subfolder = part after folder_root, blank if file is at the root level
        rel_parts = Path(rec["rel_path"].replace("\\", "/")).parts
        subfolder = rel_parts[0] if len(rel_parts) > 1 else ""

        results.append({
            "subfolder":  subfolder,
            "filename":   rec["filename"],
            "extension":  rec["extension"],
            "size_bytes": rec["size_bytes"],
            "full_path":  full_path,
            "norm_path":  norm_path,
            "in_db":      True,
            "db_id":      rec["db_id"],
        })

    results.sort(key=lambda r: (r["subfolder"].lower(), r["full_path"].lower()))
    print(f"  Using {len(results):,} files from DB (no disk walk)")
    return results


# ---------------------------------------------------------------------------
# Enrich each record with match/annotation data
# ---------------------------------------------------------------------------

def enrich(records: list[dict], db_data: dict) -> list[dict]:
    file_matches   = db_data["file_matches"]
    reel_titles    = db_data["reel_titles"]
    reel_shotlists = db_data["reel_shotlists"]
    annotations    = db_data["annotations"]

    for rec in records:
        db_id = rec.get("db_id")
        reels = file_matches.get(db_id, []) if db_id else []

        rec["matched_reels"]    = ", ".join(reels)
        rec["reel_title"]       = reel_titles.get(reels[0], "") if reels else ""
        rec["has_shotlist_pdf"] = "YES" if any(reel_shotlists.get(r, 0) for r in reels) else "NO"

        ann = annotations.get(rec["norm_path"], {})
        rec["expert_identifier"] = ann.get("expert_identifier", "")
        rec["expert_title"]      = ann.get("expert_title", "")
        rec["shotlist_pdf"]      = ann.get("shotlist_pdf", "")
        rec["notes"]             = ann.get("notes", "")
        rec["_previously_saved"] = bool(ann)

    return records


# ---------------------------------------------------------------------------
# Excel workbook builder
# ---------------------------------------------------------------------------

def _write_header(ws) -> None:
    for col_idx, (header, hint, width) in enumerate(ALL_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
        if col_idx <= len(AUTO_COLS):
            _fill(cell, FILL_HEADER_AUTO, font=FONT_HEADER)
        else:
            cell.value = f"{header}\n({hint})"
            _fill(cell, FILL_HEADER_EXPERT, font=FONT_EXPERT_H)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(NCOLS)}1"


def _write_row(ws, row_num: int, rec: dict) -> None:
    values = [
        rec["subfolder"],
        rec["filename"],
        rec["extension"],
        fmt_bytes(rec["size_bytes"]) if rec["size_bytes"] else "",
        rec["full_path"],
        "YES" if rec["in_db"] else "NO",
        rec["db_id"] if rec["db_id"] else "",
        rec["matched_reels"],
        rec["reel_title"],
        rec["has_shotlist_pdf"],
        rec["expert_identifier"],
        rec["expert_title"],
        rec["shotlist_pdf"],
        rec["notes"],
    ]

    needs_attn = not rec["matched_reels"] and not rec["expert_identifier"]

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = FONT_PATH if col_idx == COL_FULL_PATH else FONT_NORMAL

        if col_idx <= len(AUTO_COLS):
            if needs_attn:
                cell.fill = FILL_NEEDS_ATTN
            else:
                cell.fill = FILL_AUTO
        else:
            if rec["_previously_saved"] and value:
                cell.fill = FILL_EXPERT_PREFIL
            else:
                cell.fill = FILL_EXPERT_CELL


def build_workbook(
    records: list[dict],
    *,
    unmatched_only: bool = False,
) -> openpyxl.Workbook:
    if unmatched_only:
        records = [r for r in records if not r["matched_reels"] and not r["expert_identifier"]]
        print(f"  --unmatched-only: {len(records):,} files need attention")

    # Group by subfolder
    by_folder: dict[str, list[dict]] = defaultdict(list)
    for rec in records:
        by_folder[rec["subfolder"]].append(rec)

    wb = openpyxl.Workbook()

    # --- Summary sheet ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 12
    ws_sum.column_dimensions["C"].width = 12
    ws_sum.column_dimensions["D"].width = 12
    ws_sum.column_dimensions["E"].width = 18

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for c, h in enumerate(["Subfolder", "Total Files", "In DB", "Need Attn", "Annotated"], start=1):
        cell = ws_sum.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_i, folder in enumerate(sorted(by_folder.keys()), start=2):
        recs   = by_folder[folder]
        in_db  = sum(1 for r in recs if r["in_db"])
        needs  = sum(1 for r in recs if not r["matched_reels"] and not r["expert_identifier"])
        saved  = sum(1 for r in recs if r["_previously_saved"])
        ws_sum.cell(row=row_i, column=1, value=folder)
        ws_sum.cell(row=row_i, column=2, value=len(recs))
        ws_sum.cell(row=row_i, column=3, value=in_db)
        needs_cell = ws_sum.cell(row=row_i, column=4, value=needs)
        if needs:
            needs_cell.fill = FILL_NEEDS_ATTN
        ws_sum.cell(row=row_i, column=5, value=saved)

    total_row = len(by_folder) + 2
    total_font = Font(bold=True, size=10)
    ws_sum.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws_sum.cell(row=total_row, column=2, value=len(records)).font = total_font
    ws_sum.cell(row=total_row, column=3, value=sum(1 for r in records if r["in_db"])).font = total_font
    ws_sum.cell(row=total_row, column=4, value=sum(
        1 for r in records if not r["matched_reels"] and not r["expert_identifier"]
    )).font = total_font
    ws_sum.cell(row=total_row, column=5, value=sum(1 for r in records if r["_previously_saved"])).font = total_font

    ws_sum.freeze_panes = "A2"

    # --- Single sheet with all records ---
    ws_all = wb.create_sheet(title="All Files")
    _write_header(ws_all)
    for row_i, rec in enumerate(records, start=2):
        _write_row(ws_all, row_i, rec)

    print(f"  Built workbook: {len(records):,} rows in 'All Files' sheet + Summary")
    return wb


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--db",   default=DB_PATH, help=f"SQLite DB (default: {DB_PATH})")
    ap.add_argument("--root", default=None,
                    help="Root path to scope the export (default: all DB roots)")
    ap.add_argument("--out",  default="exports/expert_review.xlsx",
                    help="Output .xlsx path (default: exports/expert_review.xlsx)")
    ap.add_argument("--from-db", action="store_true",
                    help="Use DB records only — skip slow disk walk")
    ap.add_argument("--unmatched-only", action="store_true",
                    help="Only include files with no matched reel and no prior annotation")
    args = ap.parse_args()

    print(f"\nExcel Export — Disk File Catalog for Expert Review")
    print(f"  DB:   {args.db}")
    print(f"  Root: {args.root or '(all)'}")
    print(f"  Out:  {args.out}")
    print()

    print("Loading database ...")
    db_data = load_db_data(args.db, root_filter=args.root)

    if args.from_db:
        records = files_from_db(db_data["files"], root_filter=args.root)
    else:
        if not args.root:
            print("  No --root specified; using --from-db mode (pass --root for a disk walk)")
            records = files_from_db(db_data["files"], root_filter=None)
        else:
            records = walk_root(args.root, db_data["files"])

    print("\nEnriching records with match data ...")
    records = enrich(records, db_data)

    print("\nBuilding workbook ...")
    wb = build_workbook(records, unmatched_only=args.unmatched_only)

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"\nSaved: {args.out}")
    print()


if __name__ == "__main__":
    main()
