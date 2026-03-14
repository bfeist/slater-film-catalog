"""
Excel Import — Load expert annotations back into the database.

Reads a workbook previously produced by excel_export.py that an archive
expert has partially or fully filled in.  Only rows where the expert has
entered at least one value in the yellow columns are processed; blank rows
are skipped so incremental reviews don't overwrite existing annotations
with empty data.

Results are written to the  file_annotations  table in catalog.db (created
automatically on first run).  Subsequent imports are safe to re-run: each
file path is an upsert, so existing annotations are updated rather than
duplicated.

Usage (run from scripts/files_audit/)
-----
Import a filled-in workbook:
    uv run python excel_import.py exports/expert_review_filled.xlsx

Dry-run (show what would be imported without writing):
    uv run python excel_import.py exports/expert_review_filled.xlsx --dry-run

Import and print every row being processed:
    uv run python excel_import.py exports/expert_review_filled.xlsx --verbose
"""

from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

import openpyxl

DB_PATH = "../../database/catalog.db"

# Column names produced by excel_export.py (must match, case-insensitive).
EXPECTED_HEADERS = [
    "subfolder", "filename", "extension", "size", "full_path",
    "in_db", "db_id", "matched_reels", "reel_title", "has_shotlist_pdf",
    "expert_identifier", "expert_title", "shotlist_pdf", "notes",
]

EXPERT_COL_NAMES = {"expert_identifier", "expert_title", "shotlist_pdf", "notes"}

# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS file_annotations (
    full_path           TEXT PRIMARY KEY,
    file_id             INTEGER REFERENCES files_on_disk(id),
    expert_identifier   TEXT,
    expert_title        TEXT,
    shotlist_pdf        TEXT,
    notes               TEXT,
    updated_at          TEXT DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_fa_path ON file_annotations(full_path);
CREATE INDEX IF NOT EXISTS idx_fa_file ON file_annotations(file_id);
"""

UPSERT_SQL = """
INSERT INTO file_annotations
    (full_path, file_id, expert_identifier, expert_title, shotlist_pdf, notes, updated_at)
VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
ON CONFLICT(full_path) DO UPDATE SET
    file_id           = excluded.file_id,
    expert_identifier = excluded.expert_identifier,
    expert_title      = excluded.expert_title,
    shotlist_pdf      = excluded.shotlist_pdf,
    notes             = excluded.notes,
    updated_at        = excluded.updated_at
"""


def ensure_table(conn: sqlite3.Connection) -> None:
    for stmt in CREATE_TABLE_SQL.strip().split(";"):
        stmt = stmt.strip()
        if stmt:
            conn.execute(stmt)
    conn.commit()


def load_file_id_index(conn: sqlite3.Connection) -> dict[str, int]:
    """Map normalised full_path -> files_on_disk.id for fast lookups."""
    rows = conn.execute(
        "SELECT id, folder_root, rel_path FROM files_on_disk"
    ).fetchall()
    index: dict[str, int] = {}
    for row_id, folder_root, rel_path in rows:
        full = (folder_root.rstrip("/") + "/" + rel_path).replace("\\", "/").lower()
        index[full] = row_id
    return index


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def normalise_header(h: str) -> str:
    """Strip any tooltip text after newline, lowercase, strip spaces."""
    return h.split("\n")[0].strip().lower()


def parse_workbook(xlsx_path: str) -> list[dict]:
    """
    Read all data sheets (skipping 'Summary') in the workbook.
    Returns a flat list of row dicts (only rows with at least one expert value).
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    all_rows: list[dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "summary":
            continue

        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        # Parse header row
        try:
            raw_headers = next(rows_iter)
        except StopIteration:
            continue

        headers = [normalise_header(str(h)) if h else "" for h in raw_headers]

        # Build column index map
        col_map: dict[str, int] = {}
        for i, h in enumerate(headers):
            if h in EXPECTED_HEADERS:
                col_map[h] = i

        missing = [h for h in EXPECTED_HEADERS if h not in col_map]
        if missing:
            print(f"  WARNING: sheet '{sheet_name}' is missing columns: {missing}")
            continue

        sheet_rows = 0
        for raw_row in rows_iter:
            def get(col_name: str) -> str:
                idx = col_map[col_name]
                val = raw_row[idx] if idx < len(raw_row) else None
                return str(val).strip() if val is not None else ""

            # Skip rows where all expert columns are blank
            expert_values = {col: get(col) for col in EXPERT_COL_NAMES}
            if not any(expert_values.values()):
                continue

            full_path = get("full_path")
            if not full_path:
                continue

            row = {
                "sheet":      sheet_name,
                "full_path":  full_path,
                "db_id_str":  get("db_id"),
                **expert_values,
            }
            all_rows.append(row)
            sheet_rows += 1

        if sheet_rows:
            print(f"  Sheet '{sheet_name}': {sheet_rows} annotated rows")

    wb.close()
    return all_rows


# ---------------------------------------------------------------------------
# Import logic
# ---------------------------------------------------------------------------

def run_import(
    rows: list[dict],
    conn: sqlite3.Connection,
    file_id_index: dict[str, int],
    *,
    dry_run: bool = False,
    verbose: bool = False,
) -> dict:
    inserted = updated = skipped = errors = 0

    # Fetch existing annotations for comparison
    existing = {
        row[0]: row
        for row in conn.execute(
            "SELECT full_path, expert_identifier, expert_title, shotlist_pdf, notes "
            "FROM file_annotations"
        ).fetchall()
    }

    for row in rows:
        full_path = row["full_path"]
        norm_path = full_path.replace("\\", "/").lower()

        # Resolve file_id from index (prefer over any value in the spreadsheet)
        file_id = file_id_index.get(norm_path)

        expert_identifier = row["expert_identifier"] or None
        expert_title      = row["expert_title"] or None
        shotlist_pdf      = row["shotlist_pdf"] or None
        notes             = row["notes"] or None

        is_update = norm_path in existing
        action    = "UPDATE" if is_update else "INSERT"

        if verbose:
            print(
                f"  [{action}] {full_path}\n"
                f"           id={expert_identifier!r}  title={expert_title!r}  "
                f"pdf={shotlist_pdf!r}  notes={notes!r}"
            )

        if not dry_run:
            try:
                conn.execute(UPSERT_SQL, (
                    full_path, file_id,
                    expert_identifier, expert_title, shotlist_pdf, notes,
                ))
                if is_update:
                    updated += 1
                else:
                    inserted += 1
            except sqlite3.Error as exc:
                print(f"  ERROR on {full_path}: {exc}")
                errors += 1
        else:
            if is_update:
                updated += 1
            else:
                inserted += 1

    if not dry_run:
        conn.commit()

    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("xlsx", help="Path to the filled-in .xlsx file")
    ap.add_argument("--db",        default=DB_PATH, help=f"SQLite DB (default: {DB_PATH})")
    ap.add_argument("--dry-run",   action="store_true",
                    help="Show what would be imported without writing to DB")
    ap.add_argument("--verbose",   action="store_true",
                    help="Print each row as it is processed")
    args = ap.parse_args()

    if not Path(args.xlsx).exists():
        print(f"ERROR: file not found: {args.xlsx}")
        raise SystemExit(1)

    print(f"\nExcel Import — Loading expert annotations")
    print(f"  Source: {args.xlsx}")
    print(f"  DB:     {args.db}")
    if args.dry_run:
        print("  *** DRY RUN — no changes will be written ***")
    print()

    print("Parsing workbook ...")
    rows = parse_workbook(args.xlsx)
    print(f"  Total annotated rows found: {len(rows):,}")

    if not rows:
        print("  Nothing to import.")
        return

    print(f"\n{'Would write' if args.dry_run else 'Writing'} to database ...")
    conn = sqlite3.connect(args.db)

    if not args.dry_run:
        ensure_table(conn)

    file_id_index = load_file_id_index(conn)

    stats = run_import(rows, conn, file_id_index, dry_run=args.dry_run, verbose=args.verbose)
    conn.close()

    dry = " (dry run)" if args.dry_run else ""
    print(f"\nDone{dry}:")
    print(f"  Inserted: {stats['inserted']:,}")
    print(f"  Updated:  {stats['updated']:,}")
    if stats["errors"]:
        print(f"  Errors:   {stats['errors']:,}")
    print()


if __name__ == "__main__":
    main()
