"""
Stage 1b: Ingest First Steps Master Scanning List into the existing SQLite database.

Source: input_indexes/First Steps - Master Scanning List.xlsx

Sheets ingested:
    Project 1 - NARA Panavision Col   — 210 rows, 255-PV-* identifiers
    Project 1 - NARA Special Venue    — 28 rows, mixed non-255 and 255-SE/KSC identifiers
    Project 2 - 1635 NARA Selection   — 99 rows, 255-WS/PV/FR/HQ/S/SE identifiers

Schema changes applied automatically on first run:
    film_rolls:   + nara_roll_number, gauge_65mm, gauge_35mm, nara_shot_list_ref, notes
    transfers:    + reel_part
    nara_citations: new table

Identifier rule:
    NARA RG-255 identifiers carry a "255-" collection prefix in the source data.
    We always strip it before storing in film_rolls.identifier so that
    "255-FR-0001" → "FR-0001" and merges with rows from 1b_ingest_excel.py.
    Non-255 identifiers (40-UD-*, 43-US-*, 151.1-*) are stored as-is.
    When matching files on disk, re-prepend "255-" for collection members where needed.

Usage:
    uv run python scripts/1b_ingest_first_steps.py              # ingest all sheets
    uv run python scripts/1b_ingest_first_steps.py --force       # drop/re-insert NARA rows
    uv run python scripts/1b_ingest_first_steps.py --stats       # print stats only
    uv run python scripts/1b_ingest_first_steps.py --sheet pv    # pv | venue | p2
"""

import argparse
import os
import re
import sqlite3
import time
from datetime import datetime
from pathlib import Path

import openpyxl

EXCEL_PATH = "input_indexes/First Steps - Master Scanning List.xlsx"
DB_PATH = "data/01b_excel.db"

SHEET_PV = "Project 1 - NARA Panavision Col"
SHEET_VENUE = "Project 1 - NARA Special Venue "   # trailing space in actual sheet name
SHEET_P2 = "Project 2 - 1635 NARA Selection"

# ---------------------------------------------------------------------------
# Migration — add new columns / tables to existing DB
# ---------------------------------------------------------------------------

MIGRATION_SQL = """
-- New columns on film_rolls (SQLite ignores ADD COLUMN if already present when
-- wrapped in a try/except; we use a helper function below instead)

CREATE TABLE IF NOT EXISTS nara_citations (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    reel_identifier TEXT NOT NULL,
    citation        TEXT NOT NULL,
    citation_type   TEXT,            -- ksc_number | as_magazine | jsc_collection |
                                     -- hq_collection | engineering_collection |
                                     -- source_reel | other
    source_column   TEXT,
    source_sheet    TEXT
);

CREATE INDEX IF NOT EXISTS idx_nc_reel ON nara_citations(reel_identifier);
CREATE INDEX IF NOT EXISTS idx_nc_cit  ON nara_citations(citation);
"""

NEW_FILM_ROLL_COLUMNS = [
    ("nara_roll_number",  "TEXT"),    # Technicolor/Panavision roll number
    ("gauge_65mm",        "INTEGER DEFAULT 0"),
    ("gauge_35mm",        "INTEGER DEFAULT 0"),
    ("nara_shot_list_ref","TEXT"),    # raw "Shot List" column value
    ("notes",             "TEXT"),    # merged Notes / Comments / Slater notes
]

NEW_TRANSFER_COLUMNS = [
    ("reel_part", "INTEGER"),         # reel number within a multi-reel item
]


def apply_migrations(db: sqlite3.Connection) -> None:
    """Add new columns / tables to an existing DB without destroying data."""
    db.executescript(MIGRATION_SQL)

    existing_fr_cols = {row[1] for row in db.execute("PRAGMA table_info(film_rolls)")}
    for col, typedef in NEW_FILM_ROLL_COLUMNS:
        if col not in existing_fr_cols:
            db.execute(f"ALTER TABLE film_rolls ADD COLUMN {col} {typedef}")
            print(f"  [migration] film_rolls += {col}")

    existing_xfer_cols = {row[1] for row in db.execute("PRAGMA table_info(transfers)")}
    for col, typedef in NEW_TRANSFER_COLUMNS:
        if col not in existing_xfer_cols:
            db.execute(f"ALTER TABLE transfers ADD COLUMN {col} {typedef}")
            print(f"  [migration] transfers += {col}")

    db.commit()


# ---------------------------------------------------------------------------
# Identifier helpers
# ---------------------------------------------------------------------------

def normalize_nara_id(raw: str) -> str:
    """Strip '255-' prefix from NARA RG-255 identifiers; leave others unchanged."""
    s = str(raw).strip()
    if s.startswith("255-"):
        return s[4:]
    return s


def extract_id_prefix(identifier: str) -> str:
    """Normalize identifier to a prefix category (extended for NARA types)."""
    s = identifier.strip()
    for prefix in [
        "JSCmSTS", "JSCmND", "JSCm", "VJSC", "JSC",
        "FR-", "AK-", "BRF", "CMP", "CL",
        "CS-", "HQ-", "KSC", "LRL-", "SL", "ASR", "VCL", "EC",
        # NARA-specific prefixes (after 255- has been stripped)
        "PV-", "WS-", "SE-", "255-KSC",
    ]:
        if s.startswith(prefix):
            return prefix.rstrip("-")
    # Non-255 special collections
    if s.startswith("40-UD"):
        return "40-UD"
    if s.startswith("43-US"):
        return "43-US"
    if s.startswith("151."):
        return "151"
    if s.startswith("KODAK"):
        return "KODAK"
    if s.startswith("255-"):
        return "255-S"
    if re.match(r"^\d+-\d+-\d+$", s):
        return "numeric"
    if s.startswith("S") and len(s) > 1 and s[1:2].isdigit():
        return "S"
    return "other"


def classify_citation(citation: str) -> str:
    """Guess citation type from content."""
    c = citation.strip()
    if re.match(r"^KSC-\d{2}-", c):
        return "ksc_number"
    if re.match(r"^AS-\d+\b", c):
        return "as_magazine"
    if "JSC File Roll" in c or "JSC Engineering" in c:
        return "jsc_collection"
    if "HQ Stock" in c or "HQ stock" in c:
        return "hq_collection"
    if "Engineering Footage" in c:
        return "engineering_collection"
    return "other"


def parse_citations(raw: str) -> list[str]:
    """Split a citations cell into individual citation strings."""
    if not raw or str(raw).strip() in ("-", ""):
        return []
    text = str(raw).replace("_x000D_\n", "\n").replace("_x000D_", "\n")
    # Split on newlines, semicolons, or comma-space where ≥2 items look like citation codes
    parts = re.split(r"\n|;\s*", text)
    result = []
    for part in parts:
        part = part.strip().strip(",").strip()
        if part and part not in ("-",):
            result.append(part)
    return result


def safe_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(str(val))
    except (ValueError, TypeError):
        return None


def format_date(val) -> tuple[str | None, str | None]:
    if val is None:
        return None, None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d"), val.strftime("%Y-%m-%d %H:%M:%S")
    raw = str(val).strip()
    return (raw, raw) if raw else (None, None)


# ---------------------------------------------------------------------------
# Upsert helpers
# ---------------------------------------------------------------------------

UPSERT_ROLL = """
INSERT INTO film_rolls
    (identifier, id_prefix, title, date_raw, feet, nara_roll_number,
     gauge_65mm, gauge_35mm, nara_shot_list_ref, notes,
     has_shotlist_pdf, has_transfer_on_disk)
VALUES (?,?,?,?,?,?,?,?,?,?,?,0)
ON CONFLICT(identifier) DO UPDATE SET
    title             = COALESCE(EXCLUDED.title, title),
    date_raw          = COALESCE(EXCLUDED.date_raw, date_raw),
    feet              = COALESCE(EXCLUDED.feet, feet),
    nara_roll_number  = COALESCE(EXCLUDED.nara_roll_number, nara_roll_number),
    gauge_65mm        = MAX(gauge_65mm, EXCLUDED.gauge_65mm),
    gauge_35mm        = MAX(gauge_35mm, EXCLUDED.gauge_35mm),
    nara_shot_list_ref= COALESCE(EXCLUDED.nara_shot_list_ref, nara_shot_list_ref),
    notes             = COALESCE(EXCLUDED.notes, notes),
    has_shotlist_pdf  = MAX(has_shotlist_pdf, EXCLUDED.has_shotlist_pdf)
"""


def insert_transfer(db: sqlite3.Connection, identifier: str, filename: str | None,
                    reel_part: int | None, file_desc: str | None,
                    audio_file: str | None, source_sheet: str) -> bool:
    """Insert a digital_file transfer, skipping if (identifier, filename) already exists."""
    if not filename:
        return False
    exists = db.execute(
        "SELECT 1 FROM transfers WHERE reel_identifier=? AND transfer_type='digital_file' AND filename=?",
        (identifier, filename),
    ).fetchone()
    if exists:
        return False
    db.execute(
        "INSERT INTO transfers (reel_identifier, transfer_type, source_tab, "
        "filename, file_description, audio_file, reel_part) VALUES (?,?,?,?,?,?,?)",
        (identifier, "digital_file", source_sheet, filename, file_desc, audio_file, reel_part),
    )
    return True


def insert_lto(db: sqlite3.Connection, identifier: str, lto: str, source_sheet: str) -> bool:
    """Insert an LTO transfer, skipping duplicates."""
    if not lto:
        return False
    exists = db.execute(
        "SELECT 1 FROM transfers WHERE reel_identifier=? AND transfer_type='lto_copy' AND lto_number=?",
        (identifier, lto),
    ).fetchone()
    if exists:
        return False
    db.execute(
        "INSERT INTO transfers (reel_identifier, transfer_type, source_tab, lto_number) VALUES (?,?,?,?)",
        (identifier, "lto_copy", source_sheet, lto),
    )
    return True


def insert_citations(db: sqlite3.Connection, identifier: str, raw: str,
                     source_column: str, source_sheet: str) -> int:
    """Parse and insert citation records, skipping duplicates."""
    count = 0
    for cit in parse_citations(raw):
        exists = db.execute(
            "SELECT 1 FROM nara_citations WHERE reel_identifier=? AND citation=?",
            (identifier, cit),
        ).fetchone()
        if not exists:
            db.execute(
                "INSERT INTO nara_citations (reel_identifier, citation, citation_type, source_column, source_sheet) "
                "VALUES (?,?,?,?,?)",
                (identifier, cit, classify_citation(cit), source_column, source_sheet),
            )
            count += 1
    return count


# ---------------------------------------------------------------------------
# Per-sheet ingest
# ---------------------------------------------------------------------------

def ingest_panavision(ws, db: sqlite3.Connection) -> dict:
    """Ingest Project 1 - NARA Panavision Col."""
    rows_inserted = 0
    xfers_inserted = 0
    cits_inserted = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        raw_id = str(row[0]).strip()
        identifier = normalize_nara_id(raw_id)

        reel_part   = safe_int(row[1])
        digital_fn  = safe_str(row[2])
        pv_roll     = safe_str(row[4])
        title       = safe_str(row[5]) or safe_str(row[7])  # Reel Title / Content fallback Content
        citations   = safe_str(row[6])
        date_raw    = safe_str(row[8])
        shot_list   = safe_str(row[9])
        sync_sound  = safe_str(row[10])
        has_65mm    = 1 if row[11] else 0
        has_35mm    = 1 if row[12] else 0
        footage     = safe_str(row[13])
        notes       = safe_str(row[14])

        has_pdf = 1 if (shot_list and shot_list not in ("-", "")) else 0

        db.execute(UPSERT_ROLL, (
            identifier, extract_id_prefix(identifier),
            title, date_raw,
            footage, pv_roll,
            has_65mm, has_35mm,
            shot_list if has_pdf else None,
            notes,
            has_pdf,
        ))
        rows_inserted += 1

        if insert_transfer(db, identifier, digital_fn, reel_part,
                           None, sync_sound, SHEET_PV):
            xfers_inserted += 1

        if citations:
            cits_inserted += insert_citations(db, identifier, citations,
                                              "Additional Citations / Source Reels", SHEET_PV)

    db.commit()
    return {"rows": rows_inserted, "transfers": xfers_inserted, "citations": cits_inserted}


def ingest_special_venue(ws, db: sqlite3.Connection) -> dict:
    """Ingest Project 1 - NARA Special Venue."""
    # Headers: NARA Local Identifier | Digital File Name | Title | Shot List |
    #          Gauge | Format | Footage | Comments | ...
    rows_inserted = 0
    xfers_inserted = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        raw_id = str(row[0]).strip()
        identifier = normalize_nara_id(raw_id)

        digital_fn  = safe_str(row[1])
        title       = safe_str(row[2])
        shot_list   = safe_str(row[3])
        gauge_raw   = safe_str(row[4])
        fmt         = safe_str(row[5])
        footage     = safe_str(row[6])
        notes       = safe_str(row[7])

        has_65mm = 1 if gauge_raw and "65" in gauge_raw else 0
        has_35mm = 1 if gauge_raw and "35" in gauge_raw else 0
        has_pdf  = 1 if (shot_list and shot_list not in ("-", "")) else 0

        db.execute(UPSERT_ROLL, (
            identifier, extract_id_prefix(identifier),
            title, None,
            footage, None,
            has_65mm, has_35mm,
            shot_list if has_pdf else None,
            notes,
            has_pdf,
        ))
        rows_inserted += 1

        if insert_transfer(db, identifier, digital_fn, None, fmt, None, SHEET_VENUE):
            xfers_inserted += 1

    db.commit()
    return {"rows": rows_inserted, "transfers": xfers_inserted, "citations": 0}


def ingest_project2(ws, db: sqlite3.Connection) -> dict:
    """Ingest Project 2 - 1635 NARA Selection.

    Headers (0-based):
     0  NARA Local Identifier
     1  Reel Number
     2  Digital File Name
     3  Reel Title
     4  Date
     5  Additional Citations / Source Reels
     6  Footage
     7  Format
     8  Shot List
     9  Sync Sound Reference
    10  Comments
    ...
    21  Final Project Plan # 2 Scanning List  (skip)
    22  Best Available NARA Source
    23  NARA JSC File Roll Collection Reel #
    24  NARA HQ Stock Footage Collection Reel #
    25  NARA JSC Engineering Footage Collection Reel #
    26  Stephen Slater Notes on Duplication
    27  LTO / Tape No
    28  Best Statement Pictures Reference Source
    29  Confirmation of NARA Source Reels
    """
    rows_inserted = 0
    xfers_inserted = 0
    cits_inserted = 0

    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [safe_str(h) for h in headers_row]

    def col(name: str) -> int | None:
        for i, h in enumerate(headers):
            if h and name.lower() in h.lower():
                return i
        return None

    COL = {
        "id":        0,
        "reel_part": 1,
        "digital_fn":2,
        "title":     3,
        "date":      4,
        "citations": 5,
        "footage":   6,
        "format":    7,
        "shot_list": 8,
        "sync":      9,
        "comments": 10,
        "best_src":  col("Best Available NARA Source"),
        "jsc_roll":  col("JSC File Roll"),
        "hq_roll":   col("HQ Stock"),
        "eng_roll":  col("Engineering Footage"),
        "slater":    col("Stephen Slater"),
        "lto":       col("LTO / Tape"),
        "stmt_ref":  col("Statement Pictures"),
        "confirm":   col("Confirmation of NARA"),
    }

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        raw_id = str(row[0]).strip()
        identifier = normalize_nara_id(raw_id)

        reel_part  = safe_int(row[COL["reel_part"]])
        digital_fn = safe_str(row[COL["digital_fn"]])
        title      = safe_str(row[COL["title"]])
        date_raw   = safe_str(row[COL["date"]])
        citations  = safe_str(row[COL["citations"]])
        footage    = safe_str(row[COL["footage"]])
        fmt        = safe_str(row[COL["format"]])
        shot_list  = safe_str(row[COL["shot_list"]])
        sync       = safe_str(row[COL["sync"]])

        # Merge notes from multiple columns
        note_parts = []
        for note_col in ("comments", "best_src", "slater", "stmt_ref"):
            v = safe_str(row[COL[note_col]]) if COL[note_col] is not None else None
            if v:
                note_parts.append(v)
        notes = " | ".join(note_parts) if note_parts else None

        has_pdf = 1 if (shot_list and shot_list not in ("-", "")) else 0

        db.execute(UPSERT_ROLL, (
            identifier, extract_id_prefix(identifier),
            title, date_raw,
            footage, None,
            0, 0,
            shot_list if has_pdf else None,
            notes,
            has_pdf,
        ))
        rows_inserted += 1

        if insert_transfer(db, identifier, digital_fn, reel_part, fmt, sync, SHEET_P2):
            xfers_inserted += 1

        # LTO transfer
        lto_val = safe_str(row[COL["lto"]]) if COL["lto"] is not None else None
        if lto_val and insert_lto(db, identifier, lto_val, SHEET_P2):
            xfers_inserted += 1

        # Citations: primary + collection cross-refs
        if citations:
            cits_inserted += insert_citations(db, identifier, citations,
                                              "Additional Citations / Source Reels", SHEET_P2)
        for cit_col_key, col_label in [
            ("jsc_roll",  "NARA JSC File Roll Collection Reel #"),
            ("hq_roll",   "NARA HQ Stock Footage Collection Reel #"),
            ("eng_roll",  "NARA JSC Engineering Footage Collection Reel #"),
            ("confirm",   "Confirmation of NARA Source Reels"),
        ]:
            idx = COL.get(cit_col_key)
            val = safe_str(row[idx]) if idx is not None else None
            if val:
                cits_inserted += insert_citations(db, identifier, val, col_label, SHEET_P2)

    db.commit()
    return {"rows": rows_inserted, "transfers": xfers_inserted, "citations": cits_inserted}


# ---------------------------------------------------------------------------
# Stats
# ---------------------------------------------------------------------------

def print_stats(db: sqlite3.Connection) -> None:
    print("\n" + "=" * 65)
    print("FIRST STEPS INGEST — DATABASE SUMMARY")
    print("=" * 65)

    for table, label in [
        ("film_rolls",     "Film Rolls (content)"),
        ("transfers",      "Transfers (instances)"),
        ("nara_citations", "NARA Citations"),
    ]:
        try:
            count = db.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            print(f"  {label:35s}: {count:>8,d} rows")
        except sqlite3.OperationalError:
            print(f"  {label:35s}: (not found)")

    # New NARA prefixes
    print(f"\n  {'--- NARA Film Rolls by prefix ---':^45}")
    for prefix in ("PV", "WS", "SE", "HQ", "40-UD", "43-US", "151", "KODAK"):
        count = db.execute(
            "SELECT COUNT(*) FROM film_rolls WHERE id_prefix=?", (prefix,)
        ).fetchone()[0]
        if count:
            print(f"    {prefix:20s}: {count:>8,d}")

    # Citation types
    print(f"\n  {'--- NARA Citations by type ---':^45}")
    for row in db.execute(
        "SELECT citation_type, COUNT(*) FROM nara_citations GROUP BY citation_type ORDER BY 2 DESC"
    ):
        print(f"    {str(row[0]):25s}: {row[1]:>8,d}")

    # Source reels: 255-FR rows that now have nara citations
    fr_with_cit = db.execute(
        "SELECT COUNT(DISTINCT reel_identifier) FROM nara_citations "
        "WHERE reel_identifier LIKE 'FR-%'"
    ).fetchone()[0]
    print(f"\n  Existing FR-* reels enriched with citations: {fr_with_cit:>6,d}")

    # Manifest
    print(f"\n  {'--- Processing info ---':^45}")
    for row in db.execute(
        "SELECT key, value FROM _manifest WHERE key LIKE 'first_steps%' ORDER BY key"
    ):
        print(f"    {row[0]:40s}: {row[1]}")

    print("=" * 65)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Stage 1b: Ingest First Steps Master Scanning List into SQLite DB"
    )
    parser.add_argument("--force", action="store_true",
                        help="Delete all rows originating from this spreadsheet and re-insert")
    parser.add_argument("--stats", action="store_true",
                        help="Print stats only (no ingest)")
    parser.add_argument("--sheet", choices=["pv", "venue", "p2", "all"], default="all",
                        help="Which sheet(s) to ingest (default: all)")
    args = parser.parse_args()

    if not os.path.exists(DB_PATH):
        print(f"ERROR: Database not found at {DB_PATH}")
        print("Run scripts/one_time/1b_ingest_excel.py first to create the base database.")
        return

    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA synchronous=NORMAL")

    apply_migrations(db)

    if args.stats:
        print_stats(db)
        db.close()
        return

    if args.force:
        print("--force: removing rows sourced from First Steps spreadsheet...")
        db.execute("DELETE FROM nara_citations WHERE source_sheet IN (?,?,?)",
                   (SHEET_PV, SHEET_VENUE, SHEET_P2))
        db.execute("DELETE FROM transfers WHERE source_tab IN (?,?,?)",
                   (SHEET_PV, SHEET_VENUE, SHEET_P2))
        # Note: we do NOT delete film_rolls rows because they may have data from other sources.
        db.commit()
        print("  Done.")

    print(f"Loading {EXCEL_PATH}...")
    t0 = time.time()
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"  Loaded in {time.time() - t0:.1f}s")

    total_rows = total_xfers = total_cits = 0

    def run_sheet(label: str, sheet_name: str, fn):
        nonlocal total_rows, total_xfers, total_cits
        t1 = time.time()
        print(f"\nIngesting '{sheet_name}'...", end=" ", flush=True)
        ws = wb[sheet_name]
        result = fn(ws, db)
        elapsed = time.time() - t1
        print(f"{result['rows']:,d} rolls, {result['transfers']:,d} transfers, "
              f"{result.get('citations', 0):,d} citations ({elapsed:.1f}s)")
        total_rows  += result["rows"]
        total_xfers += result["transfers"]
        total_cits  += result.get("citations", 0)
        db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
                   (f"first_steps_{label}_rows", str(result["rows"])))

    if args.sheet in ("pv", "all"):
        run_sheet("pv", SHEET_PV, ingest_panavision)
    if args.sheet in ("venue", "all"):
        run_sheet("venue", SHEET_VENUE, ingest_special_venue)
    if args.sheet in ("p2", "all"):
        run_sheet("p2", SHEET_P2, ingest_project2)

    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("first_steps_total_rows",      str(total_rows)))
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("first_steps_total_transfers", str(total_xfers)))
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("first_steps_total_citations", str(total_cits)))
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("first_steps_completed_at", time.strftime("%Y-%m-%dT%H:%M:%S")))
    db.commit()

    wb.close()

    print(f"\nTotal: {total_rows:,d} rolls, {total_xfers:,d} transfers, {total_cits:,d} citations")
    print_stats(db)
    db.close()


if __name__ == "__main__":
    main()
