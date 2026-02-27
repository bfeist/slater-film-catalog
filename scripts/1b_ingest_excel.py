"""
Stage 1b: Ingest ApolloReelsMaster.xlsx into a normalised SQLite database.

Data model:
    film_rolls    — The film roll (content) definition: what was filmed, when, by whom.
                    Enriched with MOCR and Apollo 17 content metadata.
    transfers     — Known physical/digital instances of each film roll:
                    HD dubs, LTO copies, Discovery tape captures, VRDS refs, files.
                    A Discovery tape is a compilation — many film rolls per tape.
    discovery_shotlist / discovery_timecodes
                  — Per-tape shot descriptions with timecodes.
                    Each tape contains multiple FR-numbered film rolls.

Outputs:
    data/01b_excel.db  — single SQLite file

Usage:
    uv run python scripts/1b_ingest_excel.py              # full ingest
    uv run python scripts/1b_ingest_excel.py --force       # re-create from scratch
    uv run python scripts/1b_ingest_excel.py --stats       # print stats only (no ingest)
"""

import argparse
import os
import re
import sqlite3
import time
from datetime import datetime
from pathlib import Path

EXCEL_PATH = "input_indexes/ApolloReelsMaster.xlsx"
DB_PATH = "data/01b_excel.db"


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

SCHEMA_SQL = """
-- =========================================================================
-- CONTENT LAYER — what was filmed
-- =========================================================================

-- Film Rolls: one row per unique content unit (the original film roll).
-- "FR" = Film Roll in NASA archive terminology. Not all start with FR- prefix.
-- Transfer quality, format, or location are NOT properties of the film roll.
CREATE TABLE IF NOT EXISTS film_rolls (
    identifier      TEXT PRIMARY KEY,        -- FR-XXXX, AK-XXX, JSCmSTS..., etc.
    id_prefix       TEXT,                    -- Normalized prefix (FR, AK, JSC, BRF, etc.)
    title           TEXT,                    -- Canonical combined title (Concat Title)
    orig_title      TEXT,                    -- Original catalog title
    date            TEXT,                    -- ISO date (YYYY-MM-DD)
    date_raw        TEXT,                    -- Original date value
    feet            TEXT,                    -- Film length in feet
    minutes         TEXT,                    -- Duration in minutes
    audio           TEXT,                    -- Original recording audio (SOF/SIL/MOS)
    description     TEXT,                    -- Content description
    mission         TEXT,                    -- Mission name (from MOCR, if known)
    has_shotlist_pdf INTEGER DEFAULT 0,       -- 1 if matching PDF exists in shotlist folder
    has_transfer_on_disk INTEGER DEFAULT 0,   -- 1 if verified file exists on /o/ (set by Stage 1c)
    rowid_excel     INTEGER                  -- 1-based row in Master List sheet
);

CREATE INDEX IF NOT EXISTS idx_fr_prefix ON film_rolls(id_prefix);
CREATE INDEX IF NOT EXISTS idx_fr_date ON film_rolls(date);
CREATE INDEX IF NOT EXISTS idx_fr_mission ON film_rolls(mission);

-- Discovery Shot List: tape-level content descriptions (964 rows, 291 tapes).
-- Each Discovery tape is a COMPILATION containing multiple film rolls (FR-numbered).
-- The identifier column may list comma-separated FR numbers for the rolls on that tape.
CREATE TABLE IF NOT EXISTS discovery_shotlist (
    rowid           INTEGER PRIMARY KEY AUTOINCREMENT,
    identifier      TEXT,                    -- FR numbers (may be comma-separated)
    tape_number     INTEGER NOT NULL,
    description     TEXT,
    shotlist_raw    TEXT                      -- Full timecoded shotlist text
);

CREATE INDEX IF NOT EXISTS idx_disc_tape ON discovery_shotlist(tape_number);
CREATE INDEX IF NOT EXISTS idx_disc_identifier ON discovery_shotlist(identifier);

-- Parsed timecoded entries from Discovery Shot List
CREATE TABLE IF NOT EXISTS discovery_timecodes (
    tape_number     INTEGER NOT NULL,
    timecode        TEXT NOT NULL,            -- HH:MM:SS
    description     TEXT,
    parent_rowid    INTEGER REFERENCES discovery_shotlist(rowid)
);

CREATE INDEX IF NOT EXISTS idx_dtc_tape ON discovery_timecodes(tape_number);

-- =========================================================================
-- TRANSFER LAYER — known instances/copies of a reel
-- =========================================================================

-- Transfers: each known physical or digital copy of a film roll.
-- A single film roll may have 0..N transfers (LTO copies, HD dubs, etc.).
-- A Discovery tape capture is many-to-one: multiple film rolls → one tape.
CREATE TABLE IF NOT EXISTS transfers (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    reel_identifier TEXT NOT NULL REFERENCES film_rolls(identifier),
    transfer_type   TEXT NOT NULL,            -- 'lto_copy', 'hd_dub', 'discovery_capture',
                                              -- 'vrds_ref', 'digital_file'
    source_tab      TEXT,                     -- Which Excel tab this came from

    -- Location / reference
    lto_number      TEXT,                     -- L-number (e.g. L000881)
    video_file_ref  TEXT,                     -- Raw VideoFile column value
    tape_number     TEXT,                     -- HD tape # or Discovery tape #
    cut_number      INTEGER,                  -- Cut on the tape
    cut_length      TEXT,                     -- Cut duration

    -- File on disk
    filename        TEXT,                     -- Actual filename (e.g. Tape 501 - Self Contained.mov)
    file_path       TEXT,                     -- Expected path on /o/
    file_description TEXT,                    -- Format (ProRes 422 HQ 1080p, etc.)
    file_audio      TEXT,                     -- Audio info for the file
    audio_file      TEXT,                     -- Separate audio file reference

    -- Status
    transfer_status TEXT,                     -- 'Yes', etc. (HD Transfer flag)

    -- Apollo 17 transfer properties
    creator         TEXT,                     -- Who made this transfer (A17 only)
    prime_data_tape TEXT                      -- Source tape for this transfer (A17 only)
);

CREATE INDEX IF NOT EXISTS idx_xfer_reel ON transfers(reel_identifier);
CREATE INDEX IF NOT EXISTS idx_xfer_type ON transfers(transfer_type);
CREATE INDEX IF NOT EXISTS idx_xfer_lto ON transfers(lto_number);
CREATE INDEX IF NOT EXISTS idx_xfer_tape ON transfers(tape_number);

-- =========================================================================
-- METADATA
-- =========================================================================

CREATE TABLE IF NOT EXISTS _manifest (
    key             TEXT PRIMARY KEY,
    value           TEXT
);
"""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def extract_id_prefix(identifier: str) -> str:
    """Normalize identifier to a prefix category."""
    s = identifier.strip()
    for prefix in [
        "JSCmSTS", "JSCmND", "JSCm", "VJSC", "JSC",
        "FR-", "AK-", "BRF", "CMP", "CL",
        "CS-", "HQ-", "KSC", "LRL-", "SL", "ASR", "VCL", "EC",
    ]:
        if s.startswith(prefix):
            return prefix.rstrip("-")
    if s.startswith("255-"):
        return "255-S"
    if re.match(r"^\d+-\d+-\d+$", s):
        return "numeric"
    if s.startswith("S") and len(s) > 1 and s[1:2].isdigit():
        return "S"
    return "other"


def extract_l_number(video_file: str | None) -> str | None:
    """Extract L-number from VideoFile column (e.g. 'L000881/AK-001' -> 'L000881')."""
    if not video_file:
        return None
    m = re.match(r"(L\d{6})", str(video_file))
    return m.group(1) if m else None


def format_date(val) -> tuple[str | None, str | None]:
    """Convert Excel date value to ISO string. Returns (iso_date, raw_value)."""
    if val is None:
        return None, None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d"), val.strftime("%Y-%m-%d %H:%M:%S")
    raw = str(val).strip()
    if not raw:
        return None, None
    return raw, raw


def safe_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def parse_timecoded_entries(shotlist_text: str) -> list[tuple[str, str]]:
    """Extract (timecode, description) pairs from Discovery shotlist text."""
    if not shotlist_text:
        return []
    text = shotlist_text.replace("_x000D_\n", "\n").replace("_x000D_", "\n")
    entries = []
    pattern = re.compile(r"(\d{2}:\d{2}:\d{2})\s+(.*?)(?=\d{2}:\d{2}:\d{2}|\Z)", re.DOTALL)
    for m in pattern.finditer(text):
        tc = m.group(1)
        desc = re.sub(r"\s+", " ", m.group(2)).strip()
        if desc:
            entries.append((tc, desc))
    return entries


# ---------------------------------------------------------------------------
# Tape-to-file mapping
# ---------------------------------------------------------------------------

TAPE_FOLDER_RANGES = [
    (501, 562, "Master 1"),
    (563, 625, "Master 2"),
    (626, 712, "Master 3"),
    (713, 886, "Master 4"),
]


def tape_path(tape_num: int) -> tuple[str | None, str | None]:
    """Return (filename, expected_path) for a Discovery tape number."""
    for start, end, folder in TAPE_FOLDER_RANGES:
        if start <= tape_num <= end:
            fn = f"Tape {tape_num} - Self Contained.mov"
            return fn, f"/o/{folder}/{fn}"
    return None, None


# ---------------------------------------------------------------------------
# Ingest functions
# ---------------------------------------------------------------------------

def ingest_rolls_and_transfers(wb, db: sqlite3.Connection) -> tuple[int, int]:
    """
    Read the Master List and produce:
      - film_rolls (content definition)
      - transfers (each known instance/copy)

    Transfer types extracted per row (when columns populated):
      lto_copy          — MOCR LTO# or L-number in VideoFile
      vrds_ref          — VideoFile with "VRDS ITEMID:" prefix
      hd_dub            — HD TAPENO / CUT / CUTLNGTH columns
      discovery_capture — Discovery Tape # (many film rolls → one tape)
      digital_file      — Filename column (Apollo 17 files)
    """
    ws = wb["Master List"]
    roll_count = 0
    xfer_count = 0
    roll_batch = []
    xfer_batch = []

    INSERT_ROLL = "INSERT OR IGNORE INTO film_rolls VALUES (" + ",".join(["?"] * 14) + ")"
    INSERT_XFER = (
        "INSERT INTO transfers (reel_identifier, transfer_type, source_tab, "
        "lto_number, video_file_ref, tape_number, cut_number, cut_length, "
        "filename, file_path, file_description, file_audio, audio_file, "
        "transfer_status) VALUES (" + ",".join(["?"] * 14) + ")"
    )

    for rownum, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is None:
            break

        identifier = str(row[0]).strip()
        date_iso, date_raw = format_date(row[8])
        video_file = safe_str(row[9])
        l_number = extract_l_number(video_file)
        discovery_tape = safe_int(row[10])
        mocr_lto = safe_str(row[1])
        mocr_mission = safe_str(row[2])
        hd_transfer_flag = safe_str(row[14])
        hd_tapeno = safe_str(row[18])
        hd_cut = safe_int(row[19])
        hd_cutlngth = safe_str(row[20])
        filename = safe_str(row[21])
        file_desc = safe_str(row[22])
        file_audio = safe_str(row[23])
        audio_file = safe_str(row[24])

        # --- REEL (content definition) ---
        roll_batch.append((
            identifier,
            extract_id_prefix(identifier),
            safe_str(row[3]),   # title (Concat Title)
            safe_str(row[4]),   # orig_title
            date_iso,
            date_raw,
            safe_str(row[15]),  # feet
            safe_str(row[16]),  # minutes
            safe_str(row[17]),  # audio
            safe_str(row[11]),  # description
            mocr_mission,       # mission
            0,                  # has_shotlist_pdf (set later)
            0,                  # has_transfer_on_disk (set by Stage 1c)
            rownum,
        ))

        # --- TRANSFERS ---

        # LTO copy (from MOCR LTO# or L-number in VideoFile)
        if mocr_lto:
            xfer_batch.append((
                identifier, "lto_copy", "master_list",
                mocr_lto, video_file, None, None, None,
                None, None, None, None, None,
                hd_transfer_flag,
            ))
            xfer_count += 1
        elif l_number:
            xfer_batch.append((
                identifier, "lto_copy", "master_list",
                l_number, video_file, None, None, None,
                None, None, None, None, None,
                None,
            ))
            xfer_count += 1

        # VRDS reference
        if video_file and video_file.startswith("VRDS"):
            xfer_batch.append((
                identifier, "vrds_ref", "master_list",
                None, video_file, None, None, None,
                None, None, None, None, None,
                None,
            ))
            xfer_count += 1

        # HD dub
        if hd_tapeno:
            xfer_batch.append((
                identifier, "hd_dub", "master_list",
                None, None, hd_tapeno, hd_cut, hd_cutlngth,
                None, None, None, None, None,
                hd_transfer_flag,
            ))
            xfer_count += 1

        # Discovery tape capture (many film rolls → one compilation tape)
        if discovery_tape:
            fn, fp = tape_path(discovery_tape)
            xfer_batch.append((
                identifier, "discovery_capture", "master_list",
                None, None, str(discovery_tape), None, None,
                fn, fp, None, None, None,
                None,
            ))
            xfer_count += 1

        # Digital file (Apollo 17 Filename column)
        if filename:
            xfer_batch.append((
                identifier, "digital_file", "master_list",
                None, None, None, None, None,
                filename, None, file_desc, file_audio, audio_file,
                None,
            ))
            xfer_count += 1

        # Flush batches periodically
        if len(roll_batch) >= 5000:
            db.executemany(INSERT_ROLL, roll_batch)
            db.executemany(INSERT_XFER, xfer_batch)
            roll_count += len(roll_batch)
            roll_batch = []
            xfer_batch = []

    if roll_batch:
        db.executemany(INSERT_ROLL, roll_batch)
        roll_count += len(roll_batch)
    if xfer_batch:
        db.executemany(INSERT_XFER, xfer_batch)

    db.commit()
    return roll_count, xfer_count


def ingest_mocr(wb, db: sqlite3.Connection) -> tuple[int, int]:
    """Ingest MOCR tab — enrich film_rolls with content metadata, add transfers.

    MOCR has cleaner titles than the Master List (which duplicates/concatenates).
    Content columns (title, feet, minutes, audio) are merged into film_rolls.
    Transfer columns (LTO#, HD Transfer flag) go into transfers table.

    Returns (enriched_count, new_transfer_count).
    """
    ws = wb["MOCR"]
    enriched = 0
    new_xfers = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break

        ident = safe_str(row[0])
        lto = safe_str(row[2])
        hd_flag = safe_str(row[3])
        mocr_title = safe_str(row[4])
        mocr_feet = safe_int(row[5])
        mocr_minutes = safe_int(row[6])
        mocr_audio = safe_str(row[7])

        # Enrich film_rolls with MOCR content metadata.
        # MOCR title is authoritative (cleaner than Master List's concatenated version).
        # feet/minutes/audio fill NULLs only.
        db.execute(
            "UPDATE film_rolls SET "
            "  title = COALESCE(?, title), "
            "  feet  = COALESCE(feet, ?), "
            "  minutes = COALESCE(minutes, ?), "
            "  audio = COALESCE(audio, ?) "
            "WHERE identifier = ?",
            (mocr_title, str(mocr_feet) if mocr_feet else None,
             str(mocr_minutes) if mocr_minutes else None,
             mocr_audio, ident),
        )
        enriched += 1

        # Deduplicated LTO transfer
        if lto:
            existing = db.execute(
                "SELECT 1 FROM transfers WHERE reel_identifier=? AND transfer_type='lto_copy' AND lto_number=?",
                (ident, lto),
            ).fetchone()
            if not existing:
                db.execute(
                    "INSERT INTO transfers (reel_identifier, transfer_type, source_tab, "
                    "lto_number, transfer_status) VALUES (?,?,?,?,?)",
                    (ident, "lto_copy", "mocr", lto, hd_flag),
                )
                new_xfers += 1

        # Deduplicated HD flag
        if hd_flag and hd_flag.lower() in ("yes", "y"):
            existing = db.execute(
                "SELECT 1 FROM transfers WHERE reel_identifier=? AND transfer_type='hd_dub'",
                (ident,),
            ).fetchone()
            if not existing:
                db.execute(
                    "INSERT INTO transfers (reel_identifier, transfer_type, source_tab, "
                    "transfer_status) VALUES (?,?,?,?)",
                    (ident, "hd_dub", "mocr", hd_flag),
                )
                new_xfers += 1

        enriched += 0  # already counted
    db.commit()
    return enriched, new_xfers


def ingest_hd(wb, db: sqlite3.Connection) -> int:
    """Ingest HD tab into transfers (deduplicated)."""
    ws = wb["HD"]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break

        ident = safe_str(row[0])
        tape = safe_str(row[1])
        cut = safe_int(row[3])

        existing = db.execute(
            "SELECT 1 FROM transfers WHERE reel_identifier=? AND transfer_type='hd_dub' "
            "AND tape_number=? AND (cut_number=? OR (cut_number IS NULL AND ? IS NULL))",
            (ident, tape, cut, cut),
        ).fetchone()

        if not existing:
            db.execute(
                "INSERT INTO transfers (reel_identifier, transfer_type, source_tab, "
                "tape_number, cut_number, cut_length, transfer_status) "
                "VALUES (?,?,?,?,?,?,?)",
                (ident, "hd_dub", "hd", tape, cut,
                 safe_str(row[4]), safe_str(row[5])),
            )
        count += 1
    db.commit()
    return count


def ingest_apollo17(wb, db: sqlite3.Connection) -> tuple[int, int]:
    """Ingest Apollo 17 tab — enrich film_rolls with content, add transfers.

    Content columns (title, description, feet, minutes, audio) merge into film_rolls.
    Transfer columns (filename, file_description, creator, prime_data_tape) go
    into the transfers table as digital_file entries.

    Returns (enriched_count, new_transfer_count).
    """
    ws = wb["17"]
    enriched = 0
    new_xfers = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        ident = safe_str(row[0])
        if not ident:
            continue

        a17_title = safe_str(row[4])
        date_iso, _ = format_date(row[8])
        a17_desc = safe_str(row[9])
        a17_creator = safe_str(row[10])
        a17_pdt = safe_str(row[2])  # prime_data_tape
        a17_feet = safe_int(row[5])
        a17_minutes = safe_int(row[6])
        a17_audio = safe_str(row[7])

        # Enrich film_rolls with Apollo 17 content metadata.
        # Title: use A17's cleaner title (prefer over Master List's concatenated).
        # feet/minutes/audio/description: fill NULLs only.
        db.execute(
            "UPDATE film_rolls SET "
            "  title = COALESCE(?, title), "
            "  description = COALESCE(description, ?), "
            "  feet  = COALESCE(feet, ?), "
            "  minutes = COALESCE(minutes, ?), "
            "  audio = COALESCE(audio, ?) "
            "WHERE identifier = ?",
            (a17_title, a17_desc,
             str(a17_feet) if a17_feet else None,
             str(a17_minutes) if a17_minutes else None,
             a17_audio, ident),
        )
        enriched += 1

        # Digital file transfer (with creator and prime_data_tape)
        filename = safe_str(row[11])
        if filename:
            existing = db.execute(
                "SELECT 1 FROM transfers WHERE reel_identifier=? AND transfer_type='digital_file' AND filename=?",
                (ident, filename),
            ).fetchone()
            if not existing:
                db.execute(
                    "INSERT INTO transfers (reel_identifier, transfer_type, source_tab, "
                    "filename, file_description, file_audio, audio_file, transfer_status, "
                    "creator, prime_data_tape) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (ident, "digital_file", "apollo17", filename,
                     safe_str(row[12]), safe_str(row[13]), safe_str(row[14]),
                     safe_str(row[1]),
                     a17_creator, a17_pdt),
                )
                new_xfers += 1

    db.commit()
    return enriched, new_xfers


def ingest_discovery_shotlist(wb, db: sqlite3.Connection) -> tuple[int, int]:
    """Ingest DiscoveryShotList tab and parse timecoded entries.

    Each Discovery tape is a compilation containing multiple FR-numbered film rolls.
    The identifier column may list comma-separated FR numbers.
    """
    ws = wb["DiscoveryShotList"]
    row_count = 0
    tc_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        tape = row[1]
        if tape is None:
            break

        shotlist_text = safe_str(row[3])

        db.execute(
            "INSERT INTO discovery_shotlist (identifier, tape_number, description, shotlist_raw) "
            "VALUES (?,?,?,?)",
            (safe_str(row[0]), safe_int(tape), safe_str(row[2]), shotlist_text),
        )
        parent_rowid = db.execute("SELECT last_insert_rowid()").fetchone()[0]

        if shotlist_text:
            entries = parse_timecoded_entries(shotlist_text)
            for tc, desc in entries:
                db.execute(
                    "INSERT INTO discovery_timecodes VALUES (?,?,?,?)",
                    (safe_int(tape), tc, desc, parent_rowid),
                )
                tc_count += 1

        row_count += 1

    db.commit()
    return row_count, tc_count


def set_has_shotlist_pdf(db: sqlite3.Connection) -> int:
    """Mark film_rolls that have a matching PDF in the shotlist folder.

    Strips date suffixes (e.g. '2012-07-17') from PDF filenames before matching.
    Returns the number of film_rolls marked.
    """
    pdf_dir = "input_indexes/MASTER FR shotlist folder"
    if not os.path.isdir(pdf_dir):
        print(f"  Warning: PDF folder not found: {pdf_dir}")
        return 0

    pdf_stems: set[str] = set()
    for f in os.listdir(pdf_dir):
        if f.lower().endswith(".pdf"):
            stem = re.sub(r"2012-\d{2}-\d{2}", "", f.replace(".pdf", ""))
            pdf_stems.add(stem)

    if not pdf_stems:
        return 0

    count = 0
    for stem in pdf_stems:
        cur = db.execute(
            "UPDATE film_rolls SET has_shotlist_pdf = 1 WHERE identifier = ?",
            (stem,),
        )
        count += cur.rowcount

    db.commit()
    return count


# ---------------------------------------------------------------------------
# Stats
# ---------------------------------------------------------------------------

def print_stats(db: sqlite3.Connection):
    tables = [
        ("film_rolls", "Film Rolls (content)"),
        ("transfers", "Transfers (instances)"),
        ("discovery_shotlist", "Discovery Shot List"),
        ("discovery_timecodes", "Discovery Timecodes"),

    ]

    print("\n" + "=" * 65)
    print("DATABASE SUMMARY")
    print("=" * 65)

    for table, label in tables:
        try:
            count = db.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            print(f"  {label:30s}: {count:>8,d} rows")
        except sqlite3.OperationalError:
            print(f"  {label:30s}: (not found)")

    # Film rolls by prefix
    print(f"\n  {'--- Film Rolls by ID prefix ---':^40}")
    for row in db.execute(
        "SELECT id_prefix, COUNT(*) as c FROM film_rolls GROUP BY id_prefix ORDER BY c DESC"
    ):
        print(f"    {row[0]:20s}: {row[1]:>8,d}")

    # Transfers by type
    print(f"\n  {'--- Transfers by type ---':^40}")
    for row in db.execute(
        "SELECT transfer_type, COUNT(*) as c FROM transfers GROUP BY transfer_type ORDER BY c DESC"
    ):
        print(f"    {row[0]:25s}: {row[1]:>8,d}")

    # Transfers by source
    print(f"\n  {'--- Transfers by source tab ---':^40}")
    for row in db.execute(
        "SELECT source_tab, COUNT(*) as c FROM transfers GROUP BY source_tab ORDER BY c DESC"
    ):
        print(f"    {row[0]:25s}: {row[1]:>8,d}")

    # Reel coverage
    total_rolls = db.execute("SELECT COUNT(*) FROM film_rolls").fetchone()[0]
    rolls_with_xfer = db.execute(
        "SELECT COUNT(DISTINCT reel_identifier) FROM transfers"
    ).fetchone()[0]
    rolls_no_xfer = total_rolls - rolls_with_xfer
    multi = db.execute(
        "SELECT COUNT(*) FROM (SELECT reel_identifier FROM transfers GROUP BY reel_identifier HAVING COUNT(*) > 1)"
    ).fetchone()[0]
    print(f"\n  {'--- Film Roll → transfer coverage ---':^40}")
    print(f"    Rolls with >= 1 transfer:    {rolls_with_xfer:>8,d} ({100*rolls_with_xfer/total_rolls:.1f}%)")
    print(f"    Rolls with no transfer:      {rolls_no_xfer:>8,d} ({100*rolls_no_xfer/total_rolls:.1f}%)")
    print(f"    Rolls with multiple xfers:   {multi:>8,d}")

    # Discovery tape stats
    disc_tapes = db.execute("SELECT COUNT(DISTINCT tape_number) FROM discovery_shotlist").fetchone()[0]
    rolls_on_disc = db.execute(
        "SELECT COUNT(*) FROM transfers WHERE transfer_type='discovery_capture'"
    ).fetchone()[0]
    print(f"\n  {'--- Discovery tapes ---':^40}")
    print(f"    Unique compilation tapes:    {disc_tapes:>8,d}")
    print(f"    Reel captures on tapes:      {rolls_on_disc:>8,d}")
    if disc_tapes:
        print(f"    Avg rolls per tape:          {rolls_on_disc/disc_tapes:>8.1f}")

    # PDF cross-reference
    pdf_dir = "input_indexes/MASTER FR shotlist folder"
    if os.path.isdir(pdf_dir):
        pdf_stems = set()
        for f in os.listdir(pdf_dir):
            if f.lower().endswith(".pdf"):
                stem = re.sub(r"2012-\d{2}-\d{2}", "", f.replace(".pdf", ""))
                pdf_stems.add(stem)
        fr_rolls = set(
            row[0] for row in db.execute(
                "SELECT identifier FROM film_rolls WHERE id_prefix = 'FR'"
            )
        )
        matches = fr_rolls & pdf_stems
        print(f"\n  {'--- PDF cross-reference ---':^40}")
        print(f"    FR rolls in database:          {len(fr_rolls):>8,d}")
        print(f"    Unique PDF stems on disk:       {len(pdf_stems):>8,d}")
        print(f"    Overlap (FR in both):            {len(matches):>8,d}")

    # Boolean flags
    has_pdf = db.execute("SELECT COUNT(*) FROM film_rolls WHERE has_shotlist_pdf = 1").fetchone()[0]
    has_xfer = db.execute("SELECT COUNT(*) FROM film_rolls WHERE has_transfer_on_disk = 1").fetchone()[0]
    print(f"\n  {'--- Film Roll flags ---':^40}")
    print(f"    has_shotlist_pdf = 1:         {has_pdf:>8,d} ({100*has_pdf/total_rolls:.1f}%)")
    print(f"    has_transfer_on_disk = 1:     {has_xfer:>8,d} ({100*has_xfer/total_rolls:.1f}%)")
    print(f"    (transfer_on_disk set by Stage 1c directory crawl)")

    # Manifest
    print(f"\n  {'--- Processing info ---':^40}")
    for row in db.execute("SELECT key, value FROM _manifest ORDER BY key"):
        print(f"    {row[0]:30s}: {row[1]}")

    print("=" * 65)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Stage 1b: Ingest ApolloReelsMaster.xlsx into normalised SQLite"
    )
    parser.add_argument("--force", action="store_true",
                        help="Delete existing database and re-create")
    parser.add_argument("--stats", action="store_true",
                        help="Print stats only (no ingest)")
    args = parser.parse_args()

    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)

    if args.stats:
        if not os.path.exists(DB_PATH):
            print(f"Database not found: {DB_PATH}")
            return
        db = sqlite3.connect(DB_PATH)
        print_stats(db)
        db.close()
        return

    if os.path.exists(DB_PATH) and not args.force:
        db = sqlite3.connect(DB_PATH)
        try:
            existing = db.execute(
                "SELECT value FROM _manifest WHERE key = 'status'"
            ).fetchone()
            if existing and existing[0] == "complete":
                print(f"Database already exists: {DB_PATH}")
                print("Use --force to re-create, or --stats to view summary.")
                print_stats(db)
                db.close()
                return
        except sqlite3.OperationalError:
            pass
        db.close()

    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        if args.force:
            print(f"Removed existing database: {DB_PATH}")

    print(f"Loading {EXCEL_PATH}...")
    import openpyxl
    t0 = time.time()
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"  Loaded in {time.time() - t0:.1f}s — sheets: {wb.sheetnames}")

    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA synchronous=NORMAL")
    db.executescript(SCHEMA_SQL)

    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("status", "in_progress"))
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("started_at", time.strftime("%Y-%m-%dT%H:%M:%S")))
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("source_file", EXCEL_PATH))
    db.commit()

    t1 = time.time()
    print("\nIngesting Master List → film_rolls + transfers...", end=" ", flush=True)
    n_rolls, n_xfers = ingest_rolls_and_transfers(wb, db)
    print(f"{n_rolls:,d} film rolls, {n_xfers:,d} transfers ({time.time()-t1:.1f}s)")
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("rolls_from_master", str(n_rolls)))
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("transfers_from_master", str(n_xfers)))

    t1 = time.time()
    print("Enriching film_rolls from MOCR...", end=" ", flush=True)
    n_enrich, n_xfer = ingest_mocr(wb, db)
    print(f"{n_enrich:,d} enriched, {n_xfer:,d} new transfers ({time.time()-t1:.1f}s)")
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("mocr_enriched", str(n_enrich)))
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("mocr_new_transfers", str(n_xfer)))

    t1 = time.time()
    print("Ingesting HD transfers...", end=" ", flush=True)
    n = ingest_hd(wb, db)
    print(f"{n:,d} rows ({time.time()-t1:.1f}s)")
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("hd_rows", str(n)))

    t1 = time.time()
    print("Enriching film_rolls from Apollo 17...", end=" ", flush=True)
    n_enrich, n_xfer = ingest_apollo17(wb, db)
    print(f"{n_enrich:,d} enriched, {n_xfer:,d} new transfers ({time.time()-t1:.1f}s)")
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("a17_enriched", str(n_enrich)))
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("a17_new_transfers", str(n_xfer)))

    t1 = time.time()
    print("Ingesting Discovery Shot List...", end=" ", flush=True)
    n_rows, n_tc = ingest_discovery_shotlist(wb, db)
    print(f"{n_rows:,d} rows, {n_tc:,d} timecoded entries ({time.time()-t1:.1f}s)")
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("discovery_rows", str(n_rows)))
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("discovery_timecodes", str(n_tc)))

    t1 = time.time()
    print("Setting has_shotlist_pdf flags...", end=" ", flush=True)
    n_pdf = set_has_shotlist_pdf(db)
    print(f"{n_pdf:,d} film_rolls matched ({time.time()-t1:.1f}s)")
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("has_shotlist_pdf_count", str(n_pdf)))

    total_xfers = db.execute("SELECT COUNT(*) FROM transfers").fetchone()[0]
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("total_transfers", str(total_xfers)))
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("status", "complete"))
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("completed_at", time.strftime("%Y-%m-%dT%H:%M:%S")))
    db.commit()

    wb.close()

    file_size = os.path.getsize(DB_PATH)
    print(f"\nDatabase written: {DB_PATH} ({file_size/1024/1024:.1f} MB)")
    print_stats(db)
    db.close()


if __name__ == "__main__":
    main()
