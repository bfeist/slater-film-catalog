"""
Stage 1b: Ingest First Steps Master Scanning List into the existing SQLite database.

Source: input_indexes/First Steps - Master Scanning List.xlsx

Sheets ingested:
    Project 1 - NARA Panavision Col   — 210 rows, 255-PV-* identifiers
    Project 1 - NARA Special Venue    — 28 rows, mixed non-255 and 255-SE/KSC identifiers
    Project 2 - 1635 NARA Selection   — 99 rows, 255-WS/PV/FR/HQ/S/SE identifiers

Schema changes applied automatically on first run:
    film_rolls:   + nara_roll_number, film_gauge, nara_id, nara_catalog_url, notes
    transfers:    + reel_part
    nara_citations:     new table (NARA collection cross-references)
    external_file_refs: new table (S3/streaming URLs — not local files)

Identifier rule:
    NARA RG-255 identifiers carry a "255-" collection prefix in the source data.
    We always strip it before storing in film_rolls.identifier so that
    "255-FR-0001" → "FR-0001" and merges with rows from 1b_ingest_excel.py.
    Non-255 identifiers (40-UD-*, 43-US-*, 151.1-*) are stored as-is.
    When matching files on disk, re-prepend "255-" for collection members where needed.

Usage:
    uv run python scripts/1b_ingest_first_steps.py              # ingest all sheets
    uv run python scripts/1b_ingest_first_steps.py --force       # drop/re-insert NARA rows
    uv run python scripts/1b_ingest_first_steps.py --stats       # print stats only
    uv run python scripts/1b_ingest_first_steps.py --sheet pv    # pv | venue | p2
"""

import argparse
import os
import re
import sqlite3
import time
from datetime import datetime
from pathlib import Path

import openpyxl

# Resolve paths relative to project root (parent of scripts directory)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)

EXCEL_PATH = os.path.join(PROJECT_ROOT, "input_indexes/First Steps - Master Scanning List.xlsx")
DB_PATH = os.path.join(PROJECT_ROOT, "data/01b_excel.db")

SHEET_PV = "Project 1 - NARA Panavision Col"
SHEET_VENUE = "Project 1 - NARA Special Venue "   # trailing space in actual sheet name
SHEET_P2 = "Project 2 - 1635 NARA Selection"

# ---------------------------------------------------------------------------
# Migration — add new columns / tables to existing DB
# ---------------------------------------------------------------------------

MIGRATION_SQL = """
-- New columns on film_rolls (SQLite ignores ADD COLUMN if already present when
-- wrapped in a try/except; we use a helper function below instead)

CREATE TABLE IF NOT EXISTS nara_citations (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    reel_identifier TEXT NOT NULL,
    citation        TEXT NOT NULL,
    citation_type   TEXT,            -- ksc_number | as_magazine | jsc_collection |
                                     -- hq_collection | engineering_collection |
                                     -- source_reel | other
    source_column   TEXT,
    source_sheet    TEXT
);

CREATE INDEX IF NOT EXISTS idx_nc_reel ON nara_citations(reel_identifier);
CREATE INDEX IF NOT EXISTS idx_nc_cit  ON nara_citations(citation);

-- External file references: remote URLs that are NOT local disk files.
-- Examples: NARA S3 shotlist PDFs (before download), NARA streaming video.
CREATE TABLE IF NOT EXISTS external_file_refs (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    reel_identifier TEXT NOT NULL,
    url             TEXT NOT NULL,
    ref_type        TEXT,            -- 'nara_shotlist_pdf' | 'nara_streaming_video'
    filename        TEXT,            -- URL basename
    source          TEXT             -- e.g. 'nara_json'
);

CREATE INDEX IF NOT EXISTS idx_efr_reel ON external_file_refs(reel_identifier);
CREATE UNIQUE INDEX IF NOT EXISTS idx_efr_url ON external_file_refs(url);
"""

NEW_FILM_ROLL_COLUMNS = [
    ("nara_roll_number",  "TEXT"),    # Technicolor/Panavision roll number
    ("film_gauge",        "TEXT"),    # e.g. "65mm", "35mm", "16mm"
    ("nara_id",           "TEXT"),    # NARA archival description NAID
    ("nara_catalog_url",  "TEXT"),    # NARA catalog page URL
    ("notes",             "TEXT"),    # merged Notes / Comments / Slater notes
]

NEW_TRANSFER_COLUMNS = [
    ("reel_part", "INTEGER"),         # reel number within a multi-reel item
]


def apply_migrations(db: sqlite3.Connection) -> None:
    """Add new columns / tables to an existing DB without destroying data."""
    db.executescript(MIGRATION_SQL)

    existing_fr_cols = {row[1] for row in db.execute("PRAGMA table_info(film_rolls)")}
    for col, typedef in NEW_FILM_ROLL_COLUMNS:
        if col not in existing_fr_cols:
            db.execute(f"ALTER TABLE film_rolls ADD COLUMN {col} {typedef}")
            print(f"  [migration] film_rolls += {col}")

    existing_xfer_cols = {row[1] for row in db.execute("PRAGMA table_info(transfers)")}
    for col, typedef in NEW_TRANSFER_COLUMNS:
        if col not in existing_xfer_cols:
            db.execute(f"ALTER TABLE transfers ADD COLUMN {col} {typedef}")
            print(f"  [migration] transfers += {col}")

    db.commit()


# ---------------------------------------------------------------------------
# Identifier helpers
# ---------------------------------------------------------------------------

def normalize_nara_id(raw: str) -> str:
    """Strip '255-' prefix from NARA RG-255 identifiers; leave others unchanged."""
    s = str(raw).strip()
    if s.startswith("255-"):
        return s[4:]
    return s


def extract_id_prefix(identifier: str) -> str:
    """Normalize identifier to a prefix category (extended for NARA types)."""
    s = identifier.strip()
    for prefix in [
        "JSCmSTS", "JSCmND", "JSCm", "VJSC", "JSC",
        "FR-", "AK-", "BRF", "CMP", "CL",
        "CS-", "HQ-", "KSC", "LRL-", "SL", "ASR", "VCL", "EC",
        # NARA-specific prefixes (after 255- has been stripped)
        "PV-", "WS-", "SE-", "255-KSC",
    ]:
        if s.startswith(prefix):
            return prefix.rstrip("-")
    # Non-255 special collections
    if s.startswith("40-UD"):
        return "40-UD"
    if s.startswith("43-US"):
        return "43-US"
    if s.startswith("151."):
        return "151"
    if s.startswith("KODAK"):
        return "KODAK"
    if s.startswith("255-"):
        return "255-S"
    if re.match(r"^\d+-\d+-\d+$", s):
        return "numeric"
    if s.startswith("S") and len(s) > 1 and s[1:2].isdigit():
        return "S"
    return "other"


def classify_citation(citation: str) -> str:
    """Guess citation type from content."""
    c = citation.strip()
    if re.match(r"^KSC-\d{2}-", c):
        return "ksc_number"
    if re.match(r"^AS-\d+\b", c):
        return "as_magazine"
    if "JSC File Roll" in c or "JSC Engineering" in c:
        return "jsc_collection"
    if "HQ Stock" in c or "HQ stock" in c:
        return "hq_collection"
    if "Engineering Footage" in c:
        return "engineering_collection"
    return "other"


def parse_citations(raw: str) -> list[str]:
    """Split a citations cell into individual citation strings."""
    if not raw or str(raw).strip() in ("-", ""):
        return []
    text = str(raw).replace("_x000D_\n", "\n").replace("_x000D_", "\n")
    # Split on newlines, semicolons, or comma-space where ≥2 items look like citation codes
    parts = re.split(r"\n|;\s*", text)
    result = []
    for part in parts:
        part = part.strip().strip(",").strip()
        if part and part not in ("-",):
            result.append(part)
    return result


def safe_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    # openpyxl encodes \r as _x000D_ in XML-sourced cells; normalise to newline
    s = s.replace("_x000D_\n", "\n").replace("_x000D_", "\n")
    return s if s else None


def safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(str(val))
    except (ValueError, TypeError):
        return None


def format_date(val) -> tuple[str | None, str | None]:
    if val is None:
        return None, None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d"), val.strftime("%Y-%m-%d %H:%M:%S")
    raw = str(val).strip()
    return (raw, raw) if raw else (None, None)


# ---------------------------------------------------------------------------
# Upsert helpers
# ---------------------------------------------------------------------------

UPSERT_ROLL = """
INSERT INTO film_rolls
    (identifier, id_prefix, title, date_raw, feet, nara_roll_number,
     film_gauge, nara_id, nara_catalog_url,
     notes,
     has_shotlist_pdf, has_transfer_on_disk)
VALUES (?,?,?,?,?,?,?,?,?,?,?,0)
ON CONFLICT(identifier) DO UPDATE SET
    title             = COALESCE(EXCLUDED.title, title),
    date_raw          = COALESCE(EXCLUDED.date_raw, date_raw),
    feet              = COALESCE(EXCLUDED.feet, feet),
    nara_roll_number  = COALESCE(EXCLUDED.nara_roll_number, nara_roll_number),
    film_gauge        = COALESCE(EXCLUDED.film_gauge, film_gauge),
    nara_id           = COALESCE(EXCLUDED.nara_id, nara_id),
    nara_catalog_url  = COALESCE(EXCLUDED.nara_catalog_url, nara_catalog_url),
    notes             = COALESCE(EXCLUDED.notes, notes),
    has_shotlist_pdf  = MAX(has_shotlist_pdf, EXCLUDED.has_shotlist_pdf)
"""


def insert_transfer(db: sqlite3.Connection, identifier: str, filename: str | None,
                    reel_part: int | None, file_desc: str | None,
                    audio_file: str | None, source_sheet: str) -> bool:
    """Insert a digital_file transfer, skipping if (identifier, filename) already exists."""
    if not filename:
        return False
    exists = db.execute(
        "SELECT 1 FROM transfers WHERE reel_identifier=? AND transfer_type='digital_file' AND filename=?",
        (identifier, filename),
    ).fetchone()
    if exists:
        return False
    db.execute(
        "INSERT INTO transfers (reel_identifier, transfer_type, source_tab, "
        "filename, file_description, audio_file, reel_part) VALUES (?,?,?,?,?,?,?)",
        (identifier, "digital_file", source_sheet, filename, file_desc, audio_file, reel_part),
    )
    return True


def insert_lto(db: sqlite3.Connection, identifier: str, lto: str, source_sheet: str) -> bool:
    """Insert an LTO transfer, skipping duplicates."""
    if not lto:
        return False
    exists = db.execute(
        "SELECT 1 FROM transfers WHERE reel_identifier=? AND transfer_type='lto_copy' AND lto_number=?",
        (identifier, lto),
    ).fetchone()
    if exists:
        return False
    db.execute(
        "INSERT INTO transfers (reel_identifier, transfer_type, source_tab, lto_number) VALUES (?,?,?,?)",
        (identifier, "lto_copy", source_sheet, lto),
    )
    return True


def insert_citations(db: sqlite3.Connection, identifier: str, raw: str,
                     source_column: str, source_sheet: str) -> int:
    """Parse and insert citation records, skipping duplicates."""
    count = 0
    for cit in parse_citations(raw):
        exists = db.execute(
            "SELECT 1 FROM nara_citations WHERE reel_identifier=? AND citation=?",
            (identifier, cit),
        ).fetchone()
        if not exists:
            db.execute(
                "INSERT INTO nara_citations (reel_identifier, citation, citation_type, source_column, source_sheet) "
                "VALUES (?,?,?,?,?)",
                (identifier, cit, classify_citation(cit), source_column, source_sheet),
            )
            count += 1
    return count


# ---------------------------------------------------------------------------
# Per-sheet ingest
# ---------------------------------------------------------------------------

def ingest_panavision(ws, db: sqlite3.Connection) -> dict:
    """Ingest Project 1 - NARA Panavision Col."""
    rows_inserted = 0
    xfers_inserted = 0
    cits_inserted = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        raw_id = str(row[0]).strip()
        identifier = normalize_nara_id(raw_id)

        reel_part   = safe_int(row[1])
        digital_fn  = safe_str(row[2])
        pv_roll     = safe_str(row[4])
        title       = safe_str(row[5]) or safe_str(row[7])  # Reel Title / Content fallback Content
        citations   = safe_str(row[6])
        date_raw    = safe_str(row[8])
        sync_sound  = safe_str(row[10])
        has_65mm_holding = bool(row[11])  # NARA holds 65mm print
        has_35mm_holding = bool(row[12])  # NARA holds 35mm print
        footage     = safe_str(row[13])
        base_notes  = safe_str(row[14])

        # Build notes — append NARA holdings info as text
        holding_parts = []
        if has_65mm_holding:
            holding_parts.append("NARA holds 65mm print")
        if has_35mm_holding:
            holding_parts.append("NARA holds 35mm print")
        notes = " | ".join(filter(None, [base_notes] + holding_parts)) or None

        db.execute(UPSERT_ROLL, (
            identifier, extract_id_prefix(identifier),
            title, date_raw,
            footage, pv_roll,
            "65mm", None, None,   # film_gauge, nara_id, nara_catalog_url
            notes,
            0,                    # has_shotlist_pdf (Shot List column is locked Google Drive)
        ))
        rows_inserted += 1

        if insert_transfer(db, identifier, digital_fn, reel_part,
                           None, sync_sound, SHEET_PV):
            xfers_inserted += 1

        if citations:
            cits_inserted += insert_citations(db, identifier, citations,
                                              "Additional Citations / Source Reels", SHEET_PV)

    db.commit()
    return {"rows": rows_inserted, "transfers": xfers_inserted, "citations": cits_inserted}


def ingest_special_venue(ws, db: sqlite3.Connection) -> dict:
    """Ingest Project 1 - NARA Special Venue."""
    # Headers: NARA Local Identifier | Digital File Name | Title | Shot List |
    #          Gauge | Format | Footage | Comments | ...
    rows_inserted = 0
    xfers_inserted = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        raw_id = str(row[0]).strip()
        identifier = normalize_nara_id(raw_id)

        digital_fn  = safe_str(row[1])
        title       = safe_str(row[2])
        # row[3] = Shot List (locked Google Drive — not usable)
        gauge_raw   = safe_str(row[4])
        fmt         = safe_str(row[5])
        footage     = safe_str(row[6])
        notes       = safe_str(row[7])

        film_gauge = gauge_raw if gauge_raw and gauge_raw not in ("-",) else None

        db.execute(UPSERT_ROLL, (
            identifier, extract_id_prefix(identifier),
            title, None,
            footage, None,
            film_gauge, None, None,   # film_gauge, nara_id, nara_catalog_url
            notes,
            0,                        # has_shotlist_pdf
        ))
        rows_inserted += 1

        if insert_transfer(db, identifier, digital_fn, None, fmt, None, SHEET_VENUE):
            xfers_inserted += 1

    db.commit()
    return {"rows": rows_inserted, "transfers": xfers_inserted, "citations": 0}


def ingest_project2(ws, db: sqlite3.Connection) -> dict:
    """Ingest Project 2 - 1635 NARA Selection.

    Headers (0-based):
     0  NARA Local Identifier
     1  Reel Number
     2  Digital File Name
     3  Reel Title
     4  Date
     5  Additional Citations / Source Reels
     6  Footage
     7  Format
     8  Shot List
     9  Sync Sound Reference
    10  Comments
    ...
    21  Final Project Plan # 2 Scanning List  (skip)
    22  Best Available NARA Source
    23  NARA JSC File Roll Collection Reel #
    24  NARA HQ Stock Footage Collection Reel #
    25  NARA JSC Engineering Footage Collection Reel #
    26  Stephen Slater Notes on Duplication
    27  LTO / Tape No
    28  Best Statement Pictures Reference Source
    29  Confirmation of NARA Source Reels
    """
    rows_inserted = 0
    xfers_inserted = 0
    cits_inserted = 0

    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [safe_str(h) for h in headers_row]

    def col(name: str) -> int | None:
        for i, h in enumerate(headers):
            if h and name.lower() in h.lower():
                return i
        return None

    COL = {
        "id":        0,
        "reel_part": 1,
        "digital_fn":2,
        "title":     3,
        "date":      4,
        "citations": 5,
        "footage":   6,
        "format":    7,
        "shot_list": 8,
        "sync":      9,
        "comments": 10,
        "jsc_roll":  col("JSC File Roll"),
        "hq_roll":   col("HQ Stock"),
        "eng_roll":  col("Engineering Footage"),
        "slater":    col("Stephen Slater"),
    }

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        raw_id = str(row[0]).strip()
        identifier = normalize_nara_id(raw_id)

        reel_part  = safe_int(row[COL["reel_part"]])
        digital_fn = safe_str(row[COL["digital_fn"]])
        title      = safe_str(row[COL["title"]])
        date_raw   = safe_str(row[COL["date"]])
        citations  = safe_str(row[COL["citations"]])
        footage    = safe_str(row[COL["footage"]])
        fmt        = safe_str(row[COL["format"]])
        shot_list  = safe_str(row[COL["shot_list"]])
        sync       = safe_str(row[COL["sync"]])

        # Merge notes from retained columns only
        note_parts = []
        for note_col in ("comments", "slater"):
            v = safe_str(row[COL[note_col]]) if COL.get(note_col) is not None else None
            if v:
                note_parts.append(v)
        notes = " | ".join(note_parts) if note_parts else None

        # Parse film_gauge from Format column (e.g. "65mm" or "35mm")
        gauge_match = re.search(r"\b(\d+mm)\b", fmt or "", re.IGNORECASE)
        film_gauge = gauge_match.group(1).lower() if gauge_match else None

        db.execute(UPSERT_ROLL, (
            identifier, extract_id_prefix(identifier),
            title, date_raw,
            footage, None,
            film_gauge, None, None,  # film_gauge, nara_id, nara_catalog_url
            notes,
            0,                       # has_shotlist_pdf
        ))
        rows_inserted += 1

        if insert_transfer(db, identifier, digital_fn, reel_part, fmt, sync, SHEET_P2):
            xfers_inserted += 1


        # Citations: primary + collection cross-refs
        if citations:
            cits_inserted += insert_citations(db, identifier, citations,
                                              "Additional Citations / Source Reels", SHEET_P2)
        for cit_col_key, col_label in [
            ("jsc_roll",  "NARA JSC File Roll Collection Reel #"),
            ("hq_roll",   "NARA HQ Stock Footage Collection Reel #"),
            ("eng_roll",  "NARA JSC Engineering Footage Collection Reel #"),
        ]:
            idx = COL.get(cit_col_key)
            val = safe_str(row[idx]) if idx is not None else None
            if val:
                cits_inserted += insert_citations(db, identifier, val, col_label, SHEET_P2)

    db.commit()
    return {"rows": rows_inserted, "transfers": xfers_inserted, "citations": cits_inserted}


# ---------------------------------------------------------------------------
# NARA JSON ingest
# ---------------------------------------------------------------------------

NARA_JSON_PATH = os.path.join(PROJECT_ROOT, "input_indexes/nara_apollo_70mm_metadata.json")


def _parse_nara_date(dates: list[dict]) -> str | None:
    """Extract best date string from NARA dates list."""
    if not dates:
        return None
    for d in dates:
        if isinstance(d, dict):
            v = d.get("date") or d.get("dateRange", {}).get("fromDate") or d.get("dateRange", {}).get("toDate")
            if v:
                return str(v)[:10]  # YYYY-MM-DD
    return None


def ingest_nara_json(json_path: str, db: sqlite3.Connection) -> dict:
    """Ingest NARA catalog JSON into film_rolls, transfers, and nara_citations."""
    import json

    with open(json_path, encoding="utf-8") as fh:
        records = json.load(fh)

    rows_inserted = 0
    xfers_inserted = 0
    cits_inserted = 0

    for rec in records:
        local_id = str(rec.get("local_identifier", "")).strip()
        if not local_id:
            continue

        identifier = normalize_nara_id(local_id)
        naid = str(rec.get("naid", "")).strip() or None
        url  = str(rec.get("url",  "")).strip() or None

        # Title: prefer title list, fall back to description
        title = None
        raw_titles = rec.get("title", [])
        if isinstance(raw_titles, list) and raw_titles:
            title = str(raw_titles[0]).strip() or None
        elif isinstance(raw_titles, str):
            title = raw_titles.strip() or None

        description = safe_str(rec.get("description"))
        date_raw    = _parse_nara_date(rec.get("dates", []))

        # Collect digital object URLs by type
        shotlist_url = None
        mp4_url      = None
        for dobj in (rec.get("digital_objects") or []):
            obj_type = str(dobj.get("type", "")).lower()
            dl_url   = dobj.get("download_url") or dobj.get("url") or ""
            if not mp4_url and obj_type == "video":
                mp4_url = dl_url.strip() or None
            if not shotlist_url and obj_type in ("document", "pdf"):
                shotlist_url = dl_url.strip() or None

        has_shotlist = 1 if shotlist_url else 0

        db.execute(UPSERT_ROLL, (
            identifier, extract_id_prefix(identifier),
            title, date_raw,
            None, None,          # feet, nara_roll_number (set from xlsx)
            "65mm", naid, url,   # film_gauge (PV reels are 65mm), nara_id, nara_catalog_url
            description,
            has_shotlist,
        ))
        rows_inserted += 1

        # External file refs: S3 shotlist PDF and streaming video (not local files)
        for ext_url, ext_type in [
            (shotlist_url, "nara_shotlist_pdf"),
            (mp4_url,      "nara_streaming_video"),
        ]:
            if not ext_url:
                continue
            fn = ext_url.split("/")[-1].split("?")[0] or None
            exists = db.execute(
                "SELECT 1 FROM external_file_refs WHERE url=?",
                (ext_url,),
            ).fetchone()
            if not exists:
                db.execute(
                    "INSERT INTO external_file_refs (reel_identifier, url, ref_type, filename, source) "
                    "VALUES (?,?,?,?,?)",
                    (identifier, ext_url, ext_type, fn, "nara_json"),
                )
                xfers_inserted += 1

        # Citations from agency_assigned_identifiers
        for aid in (rec.get("agency_assigned_identifiers") or []):
            if not isinstance(aid, dict):
                continue
            num  = safe_str(aid.get("number"))
            note = safe_str(aid.get("organizationNote") or aid.get("note") or aid.get("type"))
            if num:
                cit_text = f"{note}: {num}" if note else num
                cit_type = "pv_roll_number" if (note and "technicolor" in note.lower()) else "other"
                exists = db.execute(
                    "SELECT 1 FROM nara_citations WHERE reel_identifier=? AND citation=?",
                    (identifier, cit_text),
                ).fetchone()
                if not exists:
                    db.execute(
                        "INSERT INTO nara_citations (reel_identifier, citation, citation_type, source_column, source_sheet) "
                        "VALUES (?,?,?,?,?)",
                        (identifier, cit_text, cit_type, "agency_assigned_identifiers", "nara_json"),
                    )
                    cits_inserted += 1

    db.commit()
    return {"rows": rows_inserted, "transfers": xfers_inserted, "citations": cits_inserted}


# ---------------------------------------------------------------------------
# Stats
# ---------------------------------------------------------------------------

def print_stats(db: sqlite3.Connection) -> None:
    print("\n" + "=" * 65)
    print("FIRST STEPS INGEST — DATABASE SUMMARY")
    print("=" * 65)

    for table, label in [
        ("film_rolls",         "Film Rolls (content)"),
        ("transfers",          "Transfers (local disk)"),
        ("nara_citations",     "NARA Citations"),
        ("external_file_refs", "External File Refs (S3/stream)"),
    ]:
        try:
            count = db.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            print(f"  {label:35s}: {count:>8,d} rows")
        except sqlite3.OperationalError:
            print(f"  {label:35s}: (not found)")

    # New NARA prefixes
    print(f"\n  {'--- NARA Film Rolls by prefix ---':^45}")
    for prefix in ("PV", "WS", "SE", "HQ", "40-UD", "43-US", "151", "KODAK"):
        count = db.execute(
            "SELECT COUNT(*) FROM film_rolls WHERE id_prefix=?", (prefix,)
        ).fetchone()[0]
        if count:
            print(f"    {prefix:20s}: {count:>8,d}")

    # Citation types
    print(f"\n  {'--- NARA Citations by type ---':^45}")
    for row in db.execute(
        "SELECT citation_type, COUNT(*) FROM nara_citations GROUP BY citation_type ORDER BY 2 DESC"
    ):
        print(f"    {str(row[0]):25s}: {row[1]:>8,d}")

    # Source reels: 255-FR rows that now have nara citations
    fr_with_cit = db.execute(
        "SELECT COUNT(DISTINCT reel_identifier) FROM nara_citations "
        "WHERE reel_identifier LIKE 'FR-%'"
    ).fetchone()[0]
    print(f"\n  Existing FR-* reels enriched with citations: {fr_with_cit:>6,d}")

    # Manifest
    print(f"\n  {'--- Processing info ---':^45}")
    for row in db.execute(
        "SELECT key, value FROM _manifest WHERE key LIKE 'first_steps%' ORDER BY key"
    ):
        print(f"    {row[0]:40s}: {row[1]}")

    print("=" * 65)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Stage 1b: Ingest First Steps Master Scanning List into SQLite DB"
    )
    parser.add_argument("--force", action="store_true",
                        help="Delete all rows originating from this spreadsheet and re-insert")
    parser.add_argument("--stats", action="store_true",
                        help="Print stats only (no ingest)")
    parser.add_argument("--source", choices=["xlsx", "nara-json", "all"], default="xlsx",
                        help="Data source to ingest (default: xlsx)")
    parser.add_argument("--sheet", choices=["pv", "venue", "p2", "all"], default="all",
                        help="Which xlsx sheet(s) to ingest (default: all)")
    parser.add_argument("--nara-json", default=NARA_JSON_PATH, metavar="PATH",
                        help="Path to NARA catalog JSON")
    args = parser.parse_args()

    if not os.path.exists(DB_PATH):
        print(f"ERROR: Database not found at {DB_PATH}")
        print("Run scripts/one_time/1b_ingest_excel.py first to create the base database.")
        return

    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA synchronous=NORMAL")

    apply_migrations(db)

    if args.stats:
        print_stats(db)
        db.close()
        return

    if args.force:
        print("--force: removing rows sourced from First Steps spreadsheet / NARA JSON...")
        sheets = [SHEET_PV, SHEET_VENUE, SHEET_P2]
        if args.source in ("nara-json", "all"):
            sheets.append("nara_json")
        placeholders = ",".join("?" * len(sheets))
        db.execute(f"DELETE FROM nara_citations WHERE source_sheet IN ({placeholders})", sheets)
        db.execute(f"DELETE FROM transfers WHERE source_tab IN ({placeholders})", sheets)
        db.execute(f"DELETE FROM external_file_refs WHERE source IN ({placeholders})", sheets)
        # Note: we do NOT delete film_rolls rows because they may have data from other sources.
        db.commit()
        print("  Done.")

    total_rows = total_xfers = total_cits = 0

    wb = None
    if args.source in ("xlsx", "all"):
        print(f"Loading {EXCEL_PATH}...")
        t0 = time.time()
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        print(f"  Loaded in {time.time() - t0:.1f}s")

        def run_sheet(label: str, sheet_name: str, fn):
            nonlocal total_rows, total_xfers, total_cits
            t1 = time.time()
            print(f"\nIngesting '{sheet_name}'...", end=" ", flush=True)
            ws = wb[sheet_name]
            result = fn(ws, db)
            elapsed = time.time() - t1
            print(f"{result['rows']:,d} rolls, {result['transfers']:,d} transfers, "
                  f"{result.get('citations', 0):,d} citations ({elapsed:.1f}s)")
            total_rows  += result["rows"]
            total_xfers += result["transfers"]
            total_cits  += result.get("citations", 0)
            db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
                       (f"first_steps_{label}_rows", str(result["rows"])))

        if args.sheet in ("pv", "all"):
            run_sheet("pv", SHEET_PV, ingest_panavision)
        if args.sheet in ("venue", "all"):
            run_sheet("venue", SHEET_VENUE, ingest_special_venue)
        if args.sheet in ("p2", "all"):
            run_sheet("p2", SHEET_P2, ingest_project2)

    if args.source in ("nara-json", "all"):
        json_path = getattr(args, "nara_json")  # argparse converts --nara-json to nara_json
        print(f"\nIngesting NARA JSON ({json_path})...", end=" ", flush=True)
        t1 = time.time()
        result = ingest_nara_json(json_path, db)
        elapsed = time.time() - t1
        print(f"{result['rows']:,d} rolls, {result['transfers']:,d} transfers, "
              f"{result.get('citations', 0):,d} citations ({elapsed:.1f}s)")
        total_rows  += result["rows"]
        total_xfers += result["transfers"]
        total_cits  += result.get("citations", 0)
        db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
                   ("first_steps_nara_json_rows", str(result["rows"])))

    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("first_steps_total_rows",      str(total_rows)))
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("first_steps_total_transfers", str(total_xfers)))
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("first_steps_total_citations", str(total_cits)))
    db.execute("INSERT OR REPLACE INTO _manifest VALUES (?,?)",
               ("first_steps_completed_at", time.strftime("%Y-%m-%dT%H:%M:%S")))
    db.commit()

    if wb is not None:
        wb.close()

    print(f"\nTotal: {total_rows:,d} rolls, {total_xfers:,d} transfers, {total_cits:,d} citations")
    print_stats(db)
    db.close()


if __name__ == "__main__":
    main()
